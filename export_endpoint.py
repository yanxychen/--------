@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        cases = data.get('cases', data.get('all_cases', [])) if data else []
        if not cases:
            return jsonify({'success': False, 'message': '没有可导出的数据'}), 400

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = '估值案例'

        headers = ['参照物位置', '土地面积 (m2)', '建筑面积 (m2)', '市场价值(万元)',
                   '建筑单价(元/m2)', '数据来源', '备注', '价格类型']
        col_widths = [60, 15, 15, 15, 18, 50, 80, 15]

        hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        hfont = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
        cfont = Font(name='Microsoft YaHei', size=10)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        calign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        lalign = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hfill; cell.font = hfont; cell.alignment = calign; cell.border = border

        for ri, case in enumerate(cases, 2):
            vals = [
                case.get('referenceLocation') or case.get('参照物位置') or '',
                case.get('landArea') or case.get('土地面积(m2)') or '不适用',
                case.get('buildingArea') or case.get('建筑面积(m2)') or 0,
                case.get('marketValue') or case.get('市场价值(万元)') or 0,
                case.get('unitPrice') or case.get('建筑单价(元/m2)') or 0,
                case.get('link') or case.get('source') or case.get('数据来源') or '',
                case.get('remark') or case.get('备注') or '',
                case.get('priceType') or case.get('价格类型') or '普通司法拍卖',
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = cfont; cell.border = border
                cell.alignment = lalign if ci in (1, 6, 7) else calign
                if ci in (3, 4, 5):
                    try:
                        fv = float(v) if v not in ('-', '不适用', '', None) else 0
                        cell.value = fv
                        cell.number_format = '#,##0.00'
                    except: cell.number_format = '#,##0.00'

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=npl_valuation_cases.xlsx'})
    except Exception as e:
        print(f'导出异常: {e}')
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500