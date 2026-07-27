#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
不良资产估值参考案例搜索工具 - 后端API (Render兼容版)

适配前端 searchService.ts 的接口契约：
- POST /api/search → {status, top3, all_cases, self_auction_count, total_count}
- GET /api/health → {status, timestamp}
"""

import os
import sys
import json
import re
import time
from datetime import datetime
from flask import Flask, request, jsonify, Response

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

TAOBAO_CHROME_FETCHER_DIR = r"D:\宝宝课程资料"
if os.path.exists(TAOBAO_CHROME_FETCHER_DIR) and TAOBAO_CHROME_FETCHER_DIR not in sys.path:
    sys.path.insert(0, TAOBAO_CHROME_FETCHER_DIR)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# CORS 支持 - 允许浏览器直接调用后端
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, ngrok-skip-browser-warning'
    response.headers['Access-Control-Max-Age'] = '86400'
    return response

@app.route('/api/search', methods=['OPTIONS'])
@app.route('/api/valuate', methods=['OPTIONS'])
@app.route('/api/health', methods=['OPTIONS'])
@app.route('/api/export', methods=['OPTIONS'])
def cors_preflight():
    resp = app.make_response(('', 204))
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, ngrok-skip-browser-warning'
    resp.headers['Access-Control-Max-Age'] = '86400'
    return resp

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)


def map_raw_to_v1(raw_item, platform, index):
    """
    将原始搜索结果映射为前端 convertPythonCaseToCase 能识别的 V1 格式字段名。
    前端 converter 支持多种字段名变体，这里同时提供 V1 中文名和英文名。
    """
    title = raw_item.get('title', '')
    address = raw_item.get('address', '')
    link = raw_item.get('link', '')
    item_id = raw_item.get('item_id', '')
    source_text = '京东拍卖' if platform == 'jd' else '淘宝司法拍卖'
    
    # ★ 链接修复：无效链接→生成平台搜索链接
    if not link or link.endswith('?id=') or link.endswith('?id') or link == '':
        title = raw_item.get('title', '')
        search_kw = title[:30] if title else raw_item.get('address', '')[:30]
        if search_kw:
            from urllib.parse import quote_plus
            encoded = quote_plus(search_kw)
            if platform == 'jd':
                link = f'https://auction.jd.com/s/list.html?keyword={encoded}'
            else:
                link = f'https://sf.taobao.com/item/list.htm?q={encoded}'
        elif item_id:
            link = f'https://sf.taobao.com/sf_item/{item_id}.htm'

    # 价格处理
    current_price = raw_item.get('current_price', '')
    start_price = raw_item.get('start_price', '')

    # 尝试从价格字符串提取数值
    def parse_price_str(p):
        if not p:
            return 0.0
        try:
            # 去掉单位文字
            s = str(p).replace('万元', '').replace('万', '').replace('元', '')
            s = s.replace(',', '').replace('，', '').strip()
            val = float(s)
            if '万' in str(p):
                val *= 10000
            return val
        except:
            return 0.0

    price_val = parse_price_str(current_price)
    start_price_val = parse_price_str(start_price)

    # 面积
    # 优先从 title 中提取（很多标题直接包含"建筑面积XX㎡"）
    title = raw_item.get('title', '')
    building_area = 0.0
    title_area = re.search(r'建筑面积[：:]?\s*([\d,]+\.?\d*)\s*[㎡平方米]', title)
    if not title_area:
        title_area = re.search(r'([\d,]+\.?\d*)\s*[㎡平方米]', title)
    if title_area:
        try:
            building_area = float(title_area.group(1).replace(',', ''))
        except:
            pass
    # 如果 title 没有，从 area 字段提取
    if not building_area:
        area_str = raw_item.get('area', '')
        if area_str:
            try:
                s = str(area_str).replace('㎡', '').replace('平方米', '').replace('平', '').replace(',', '').strip()
                building_area = float(s)
            except:
                pass

    # 单价（元/㎡） - market_value已经是元，直接÷面积
    unit_price = 0.0
    market_value_for_price = price_val if price_val > 0 else start_price_val
    if market_value_for_price > 0 and building_area > 0:
        unit_price = market_value_for_price / building_area

    # 市场价值（万元） - 用起拍价或当前价估算
    market_value_wan = 0.0
    if price_val > 0:
        market_value_wan = price_val / 10000  # 转万元
    elif start_price_val > 0:
        market_value_wan = start_price_val / 10000

    # 参照物位置：用标题或地址
    ref_location = title if title else address

    # 备注 - 遵循V1标准格式
    remark_parts = []
    
    # 拍卖轮次（一拍/二拍/变卖）
    stage = raw_item.get('current_stage', '') or raw_item.get('stage', '') or '一拍'
    
    # 日期格式：YYYY年MM月DD日
    start_date = raw_item.get('start_date', '') or ''
    date_str = ''
    if start_date and len(start_date) >= 10:
        try:
            dt = datetime.strptime(start_date[:10].replace('T', ' '), '%Y-%m-%d')
            date_str = f"{dt.year}年{dt.month:02d}月{dt.day:02d}日"
        except:
            date_str = start_date[:10]
    
    # 状态
    status = raw_item.get('status', '') or raw_item.get('statusDesc', '') or ''
    
    # 组装备注：V1标准格式
    # 一拍：2026年07月24日，起拍价：5,000,000元，状态：即将开始
    remark = ''
    if start_price_val > 0:
        parts = []
        # 轮次
        stage_part = f"{stage}：" if stage else ""
        parts.append(stage_part)
        # 日期
        if date_str:
            parts.append(f"{date_str}，")
        # 起拍价
        parts.append(f"起拍价：{start_price_val:,.0f}元")
        # 成交价（如果有）
        deal_price_raw = raw_item.get('deal_price', 0)
        try: deal_val = float(str(deal_price_raw).replace(',', ''))
        except: deal_val = 0.0
        if deal_val > 0:
            parts.append(f"，成交价：{deal_val:,.0f}元")
        # 状态
        if status:
            parts.append(f"，状态：{status}")
        remark = ''.join(parts)
    else:
        remark = ''
    
    # 追加距离信息（独立一行）
    distance_km = raw_item.get('distance_km', 0)
    if distance_km and distance_km > 0:
        remark += f"\n距离抵押物约{distance_km:.1f}公里"

    # 拍卖轮次
    auction_records = []
    stage = raw_item.get('current_stage', '') or raw_item.get('stage', '')
    status = raw_item.get('status', '')
    if stage or status:
        auction_records.append({
            'round': stage or '一拍',
            'date': '',
            'startPrice': start_price_val,
            'endPrice': price_val if status in ['已成交', '成交成功'] else 0,
            'status': status or '未知',
        })

    # 构建返回对象 - 同时提供 V1 中文字段名和英文字段名
    # 前端 convertPythonCaseToCase 会按优先级尝试各种字段名
    v1_case = {
        # V1 中文格式字段（前端 converter 的主要识别格式）
        '参照物位置': ref_location,
        '建筑面积(m²)': str(building_area) if building_area > 0 else '不适用',
        '土地面积(m²)': '不适用',
        '市场价值(万元)': str(round(market_value_wan, 2)) if market_value_wan > 0 else '不适用',
        '建筑单价(元/㎡)': str(round(unit_price, 2)) if unit_price > 0 else '不适用',
        '数据来源': link,  # ★ 完整URL（V1格式：数据来源存的是URL）
        '数据来源_链接': link,
        '备注': remark,
        '价格类型': '普通司法拍卖',

        # 英文字段名（前端 converter 的备用识别格式）
        'referenceLocation': ref_location,
        'buildingArea': building_area,
        'building_area': building_area,
        'land_area': 0,
        'marketValue': market_value_wan,
        'market_value': market_value_wan,
        'unitPrice': unit_price,
        'unit_price': unit_price,
        'address': address,
        'link': link,
        'source': link,
        'source_link': link,
        'source_text': source_text,
        'remark': remark,
        'priceType': '普通司法拍卖',
        'price_type': '普通司法拍卖',
        'item_id': item_id,

        # 拍卖记录
        'auctionRecords': auction_records,
        'auction_records': auction_records,

        # 元信息
        'platform': platform,
        'is_self_auction': False,
        'distance_km': raw_item.get('distance_km', None),
        'price_anomaly': None,
    }

    # 如果有详情数据（来自 get_taobao_detail），覆盖上面字段
    detail = raw_item.get('detail', {})
    if detail and detail.get('success'):
        if detail.get('building_area', 0) > 0:
            v1_case['建筑面积(m²)'] = str(detail['building_area'])
            v1_case['buildingArea'] = detail['building_area']
            v1_case['building_area'] = detail['building_area']
        if detail.get('address'):
            v1_case['address'] = detail['address']
            # 更新参照物位置
            if not title:
                v1_case['参照物位置'] = detail['address']
                v1_case['referenceLocation'] = detail['address']
        if detail.get('deal_price', 0) > 0:
            v1_case['市场价值(万元)'] = str(round(detail['deal_price'] / 10000, 2))
            v1_case['marketValue'] = detail['deal_price'] / 10000
            v1_case['market_value'] = detail['deal_price'] / 10000
        elif detail.get('start_price', 0) > 0:
            v1_case['市场价值(万元)'] = str(round(detail['start_price'] / 10000, 2))
            v1_case['marketValue'] = detail['start_price'] / 10000
            v1_case['market_value'] = detail['start_price'] / 10000
        # 单价
        if detail.get('deal_price', 0) > 0 and detail.get('building_area', 0) > 0:
            v1_case['建筑单价(元/㎡)'] = str(round(detail['deal_price'] / detail['building_area'], 2))
            v1_case['unitPrice'] = detail['deal_price'] / detail['building_area']
            v1_case['unit_price'] = detail['deal_price'] / detail['building_area']
        elif detail.get('start_price', 0) > 0 and detail.get('building_area', 0) > 0:
            v1_case['建筑单价(元/㎡)'] = str(round(detail['start_price'] / detail['building_area'], 2))
            v1_case['unitPrice'] = detail['start_price'] / detail['building_area']
            v1_case['unit_price'] = detail['start_price'] / detail['building_area']
        # 状态
        if detail.get('current_stage'):
            v1_case['auctionRecords'] = [{
                'round': detail['current_stage'],
                'date': str(detail.get('start_date', '')) if detail.get('start_date') else '',
                'startPrice': detail.get('start_price', 0),
                'endPrice': detail.get('deal_price', 0) if detail.get('status') == '已成交' else 0,
                'status': detail.get('status', '未知'),
            }]
            v1_case['auction_records'] = v1_case['auctionRecords']
        # 备注（用上面已格式化的remark，补充评估价/成交价）
        detail_remark_extras = []
        if detail.get('consult_price', 0) > 0:
            detail_remark_extras.append(f"评估价：{detail['consult_price']:,.0f}元")
        if detail.get('deal_price', 0) > 0 and detail.get('status') == '已成交':
            detail_remark_extras.append(f"成交价：{detail['deal_price']:,.0f}元")
        if detail_remark_extras:
            extras_str = '；'.join(detail_remark_extras)
            # 在距离信息之前插入（如果存在距离）
            dist_marker = '\n距离抵押物约'
            if dist_marker in remark:
                remark = remark.replace(dist_marker, f'；{extras_str}\n距离抵押物约')
            else:
                remark += f'；{extras_str}'
        if detail.get('status') and detail.get('status') not in str(remark):
            remark = remark.replace('\n', f'，状态：{detail["status"]}\n', 1) if '\n' in remark else remark + f'，状态：{detail["status"]}'
        v1_case['备注'] = remark
        v1_case['remark'] = remark

    return v1_case


def fetch_detail_for_item(item, platform):
    """为单个搜索结果抓取详情页数据（淘宝用MTOP API，京东用已有数据）"""
    if platform == 'taobao' and item.get('item_id'):
        try:
            from taobao_mtop_api import get_taobao_detail_mtop
            detail = get_taobao_detail_mtop(item['item_id'])
            if detail.get('success'):
                item['detail'] = detail
                # 回填所有字段
                if detail.get('start_price', 0) > 0:
                    item['start_price'] = str(detail['start_price'])
                if detail.get('consult_price', 0) > 0:
                    item['consult_price'] = str(detail['consult_price'])
                if detail.get('deal_price', 0) > 0:
                    item['deal_price'] = str(detail['deal_price'])
                if detail.get('start_date'):
                    item['start_date'] = detail['start_date']
                if detail.get('current_stage'):
                    item['current_stage'] = detail['current_stage']
                if detail.get('status'):
                    item['status'] = detail['status']
                if detail.get('address'):
                    item['address'] = detail['address']
            else:
                item['detail'] = {'success': False, 'error': detail.get('error', '')}
        except Exception as e:
            print(f"MTOP详情抓取失败 {item.get('item_id')}: {e}")
            item['detail'] = {'success': False, 'error': str(e)}
    return item


def _expand_keywords(address: str) -> list:
    """将抵押物地址拆分成由近到远的分层关键词
    第1层: 完整地址（最精确）
    第2层: 小区/楼盘名（去掉门牌号/栋号）
    第3层: 路段（XX路/XX街/XX大道）
    第4层: 行政区（XX区/XX县）
    第5层: 城市（XX市）——兜底
    """
    addr = address.strip()
    import re
    result = [addr]  # 第1层：完整地址

    # 第2层：小区/楼盘名（去掉门牌号）
    # 尝试按分隔符截取前段
    for sep in [' ', '，', ',', '、']:
        parts = addr.split(sep)
        if len(parts) > 1:
            main_part = parts[0].strip()
            if len(main_part) >= 2 and main_part not in result:
                result.append(main_part)
            break

    # 第3层：路段（XX路/XX街/XX大道）
    road_match = re.search(r'([\u4e00-\u9fff]+?(?:路|街|大道))', addr)
    if road_match:
        road = road_match.group(1)
        if road not in result:
            result.append(road)

    # 第4层：行政区——从右往左找区/县名
    # 先按市/州拆分，在最后一个分段里找区/县
    segments = re.split(r'[市州]', addr)
    area = ''
    for seg in reversed(segments):
        if not seg: continue
        m = re.search(r'([\u4e00-\u9fff]{2,3}(?:区|县))', seg)
        if m:
            area = m.group(1)
            break
    if area and area not in result:
        result.append(area)

    # 第5层：城市（XX市/XX州）——兜底
    city_match = re.search(r'(.+?[市州])', addr)
    if city_match:
        city = city_match.group(1)
        if city not in result:
            result.append(city)

    return result


def _amap_geocode(address: str) -> tuple:
    """高德地理编码：地址→(lng, lat)，失败返回None"""
    import urllib.request, urllib.parse, json
    key = os.environ.get('AMAP_API_KEY', 'd7d06a2c20dacd8c861173b82cf70d71')
    url = f"https://restapi.amap.com/v3/geocode/geo?key={key}&address={urllib.parse.quote(address[:50])}"
    try:
        data = json.loads(urllib.request.urlopen(url, timeout=5).read())
        if data.get('status') == '1' and data.get('geocodes'):
            loc = data['geocodes'][0].get('location', '')
            if loc and ',' in loc:
                parts = loc.split(',')
                return (float(parts[0]), float(parts[1]))
    except: pass
    return None


def _nearby_poi_keywords(location: tuple, property_type: str = '商业', radius: int = 500) -> list:
    """按物业类型搜索周边POI名称（类型过滤，只查2组类型码）"""
    import urllib.request, urllib.parse, json, re
    key = os.environ.get('AMAP_API_KEY', 'd7d06a2c20dacd8c861173b82cf70d71')
    loc_str = f"{location[0]},{location[1]}"
    
    # 按物业类型选择高德分类码（只查2组，快）
    type_codes = {
        '住宅': ['120300'],                    # 住宅小区
        '商业': ['120200', '060100'],           # 写字楼+购物中心
        '工业': ['141200', '140100'],           # 工业大厦+产业园区
    }
    codes = type_codes.get(property_type, type_codes['商业'])
    
    names = []
    seen = set()
    
    for code in codes:
        try:
            url = f"https://restapi.amap.com/v3/place/around?key={key}&location={loc_str}&types={code}&radius={radius}&offset=15"
            data = json.loads(urllib.request.urlopen(url, timeout=5).read())
            for p in data.get('pois', []):
                name = p.get('name', '').strip()
                if name and len(name) >= 2:
                    clean = re.sub(r'[（(][^)）]*[)）]', '', name).strip()
                    if clean and clean not in seen and len(clean) >= 4:
                        seen.add(clean)
                        names.append(clean)
        except: pass
    
    return names


def _expand_keywords_b(address: str, property_type: str = '商业') -> list:
    """高德POI关键词（限10个，过滤无效名称）"""
    loc = _amap_geocode(address)
    if not loc:
        return [address]
    
    type_map = {'residential': '住宅', 'commercial': '商业', '住宅': '住宅', '商业': '商业', '工业': '工业'}
    mapped_type = type_map.get(property_type, '商业')
    
    # 住宅/商业3km，工业10km
    radius = 10.0 if mapped_type == '工业' else 3.0
    pois = _nearby_poi_keywords(loc, mapped_type, radius)
    
    # 过滤无意义关键词（外卖柜/收货区/太短的）
    useless = ['外卖', '收货', '出入口', '垃圾站', '保安亭', '电梯', '楼梯', '通道']
    keywords = [address]
    for p in pois:
        if len(keywords) >= 20:
            break
        if any(u in p for u in useless):
            continue
        if len(p) < 3:  # 太短的无用（如 A座、D座）
            continue
        # 方案A: 地址前缀 + 新关键词 → 限定本地区域搜索
        combined = f"{address} {p}"
        if combined not in keywords:
            keywords.append(combined)
        elif p not in keywords:
            keywords.append(p)
    return keywords


def _search_with_keywords(keywords: list) -> list:
    """用关键词列表搜索淘宝拍卖案例，去重合并（限时28秒）"""
    from asset_search_api import UnifiedAuctionSearcher
    import time
    searcher = UnifiedAuctionSearcher()
    all_raw = []
    seen_links = set()
    deadline = time.time() + 28
    for kw in keywords:
        if time.time() > deadline:
            break
        try:
            items = searcher.search_all(kw, platforms=['taobao'])  # 仅淘宝
            for item in items:
                link = item.get('link', '')
                if link and link not in seen_links:
                    seen_links.add(link)
                    all_raw.append(item)
        except Exception as e:
            print(f"关键词 '{kw}' 搜索失败: {e}")
    searcher.cleanup()
    return all_raw


def _search_items(address: str, property_type: str = '商业') -> list:
    """搜索入口：高德POI关键词→搜淘宝→去重→限时"""
    return _search_with_keywords(_expand_keywords_b(address, property_type))


def _search_items_b(address: str) -> list:
    """搜索入口：策略B（高德API周边POI关键词）"""
    return _search_with_keywords(_expand_keywords_b(address))


@app.route('/api/abtest', methods=['POST'])
def ab_test():
    """A/B测试：比较两种搜索策略的结果"""
    from datetime import datetime
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': '缺少参数'}), 400
    address = data.get('address', '')
    asset_type = data.get('asset_type', '')

    # 执行A策略
    t0 = datetime.now()
    items_a = _search_items(address)
    ta = (datetime.now() - t0).total_seconds()
    filtered_a = _filter_items(_dedup_items(items_a))

    # 执行B策略
    t0 = datetime.now()
    items_b = _search_items_b(address)
    tb = (datetime.now() - t0).total_seconds()
    filtered_b = _filter_items(_dedup_items(items_b))

    # 对比分析
    links_a = {i.get('link','') for i in filtered_a}
    links_b = {i.get('link','') for i in filtered_b}
    overlap = links_a & links_b
    only_a = links_a - links_b
    only_b = links_b - links_a

    return jsonify({
        'address': address,
        'strategy_a': {
            'name': '分层地址关键词',
            'keywords': _expand_keywords(address),
            'total': len(items_a),
            'filtered': len(filtered_a),
            'time_sec': round(ta, 1),
            'sample_titles': [i.get('title','')[:40] for i in filtered_a[:5]],
        },
        'strategy_b': {
            'name': '高德API周边POI',
            'keywords': _expand_keywords_b(address),
            'total': len(items_b),
            'filtered': len(filtered_b),
            'time_sec': round(tb, 1),
            'sample_titles': [i.get('title','')[:40] for i in filtered_b[:5]],
        },
        'comparison': {
            'overlap': len(overlap),
            'unique_to_a': len(only_a),
            'unique_to_b': len(only_b),
            'total_unique': len(filtered_a) + len(filtered_b) - len(overlap),
        }
    })


def _estimate_distance(collateral: str, case_addr: str, col_coords_cache=None) -> float:
    """估算距离（公里）：高德geocode坐标 + Haversine直线距离（不用文本猜）"""
    if not case_addr or not collateral:
        return -1
    import re, json, urllib.request, urllib.parse, math

    amap_key = os.environ.get('AMAP_API_KEY', 'd7d06a2c20dacd8c861173b82cf70d71')

    def geocode(addr: str):
        try:
            url = (
                f"https://restapi.amap.com/v3/geocode/geo"
                f"?key={amap_key}&address={urllib.parse.quote(addr[:100])}"
            )
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            data = json.loads(urllib.request.urlopen(req, timeout=3).read())
            geocodes = data.get('geocodes', [])
            if geocodes:
                loc = geocodes[0].get('location', '').split(',')
                return (float(loc[0]), float(loc[1]))
        except Exception:
            pass
        return None

    def haversine(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
        R = 6371.0
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        d_phi = math.radians(lat2 - lat1)
        d_lambda = math.radians(lon2 - lon1)
        a = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        return round(R * c, 1)

    # 1. 抵押物坐标：优先用调用方传入的缓存
    col_coords = col_coords_cache or geocode(collateral)
    if not col_coords:
        return -1

    # 2. 案例坐标
    case_coords = geocode(case_addr)
    if not case_coords:
        return -1

    # 3. 返回直线距离 km
    return haversine(col_coords[0], col_coords[1], case_coords[0], case_coords[1])


def _filter_by_distance_time(items: list, address: str, property_type: str = '商业') -> list:
    """严格过滤：距离+时间"""
    try:
        from datetime import datetime
        now = datetime.now()
        is_industrial = property_type in ('工业', '土地', 'other')
        max_dist = 10.0 if is_industrial else 3.0
        max_days = 365

        # 缓存抵押物坐标只查一次
        col_coords = None
        import re as _re, json, urllib.request, urllib.parse
        amap_key = os.environ.get('AMAP_API_KEY', 'd7d06a2c20dacd8c861173b82cf70d71')
        try:
            geo_url = f"https://restapi.amap.com/v3/geocode/geo?key={amap_key}&address={urllib.parse.quote(address[:100])}"
            req = urllib.request.Request(geo_url, headers={'User-Agent': 'Mozilla/5.0'})
            resp = json.loads(urllib.request.urlopen(req, timeout=3).read())
            geocodes = resp.get('geocodes', [])
            if geocodes:
                loc = geocodes[0].get('location', '').split(',')
                col_coords = (float(loc[0]), float(loc[1]))
        except Exception:
            pass

        result = []
        for item in items:
            # 用上游算好的距离或现场算
            dist = item.get('distance_km', -1)
            if dist == -1:
                title = str(item.get('title', '') or item.get('参照物位置', '') or '')
                case_addr = str(item.get('address', '') or '')
                try:
                    dist = _estimate_distance(address, case_addr or title, col_coords)
                except Exception:
                    dist = -1
                item['distance_km'] = dist

            # 日期解析：直接从备注/标题正则提取（不依赖MTOP）
            days_old = 9999
            raw = (item.get('remark', '') or item.get('title', '') or '')
            try:
                m = re.search(r'(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})', raw)
                if m:
                    dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    days_old = (now - dt).days
            except Exception:
                pass

            # 严格时间过滤：能解析出日期的且超过1年的直接丢弃
            if days_old != 9999 and days_old > max_days:
                continue

            if dist != -1 and dist > max_dist:
                continue

            result.append(item)

        result.sort(key=lambda x: x.get('distance_km', 999) if x.get('distance_km', -1) != -1 else 999)
        return result
    except Exception as e:
        print(f"[过滤异常] {e}")
        import traceback
        traceback.print_exc()
        return items


def _dedup_items(items: list) -> list:
    """按链接去重"""
    seen = set()
    result = []
    for item in items:
        link = item.get('link', '')
        if link not in seen:
            seen.add(link)
            result.append(item)
    return result


def _filter_items(items: list) -> list:
    """排除明显不相关的非房产拍卖结果"""
    exclude_kw = ['公开选聘', '审计机构', '破产清算', '招募公告', '租赁权公告', '服务采购']
    kept = []
    for item in items:
        title = (item.get('title', '') or '')
        if any(kw in title for kw in exclude_kw):
            print(f"  过滤不相关: {title[:40]}")
            continue
        kept.append(item)
    return kept


def _enrich_details(items: list, max_items: int = 20):
    """使用本机已登录 Chrome 抓取淘宝详情页（最多max_items条）"""
    taobao_items = [i for i in items if i.get('platform') == 'taobao']
    fetch_limit = min(max_items, len(taobao_items))
    if fetch_limit == 0:
        return
    for i in range(fetch_limit):
        fetch_detail_for_item(taobao_items[i], 'taobao')


def _format_and_sort(raw_items: list, address: str) -> dict:
    """转V1格式、排序、检测自身拍卖，返回前端期望格式"""
    v1_cases = []
    for idx, raw in enumerate(raw_items):
        platform = raw.get('platform', 'unknown')
        v1_case = map_raw_to_v1(raw, platform, idx)
        v1_cases.append(v1_case)

    v1_cases.sort(key=lambda c: (
        -1 if isinstance(c.get('detail'), dict) and c['detail'].get('success') else 0,
        -1 if (c.get('buildingArea', 0) or c.get('building_area', 0) or 0) > 0 else 0,
        -1 if (c.get('marketValue', 0) or c.get('market_value', 0) or 0) > 0 else 0,
    ))

    top3 = v1_cases[:3]
    self_auction_ids = set()
    for case in v1_cases:
        case_addr = case.get('address', '') or ''
        if address in case_addr and len(address) > 5:
            case['is_self_auction'] = True
            self_auction_ids.add(case.get('item_id', ''))

    return {
        'status': 'success',
        'top3': top3,
        'all_cases': v1_cases,
        'self_auction_count': len(self_auction_ids),
        'total_count': len(v1_cases),
    }


def run_search(address, property_type=None, area=None):
    """执行搜索并返回前端期望格式"""
    try:
        all_raw = _search_items(address, property_type or '商业')
        unique_raw = _dedup_items(all_raw)
        filtered_raw = _filter_items(unique_raw)
        
        # 先MTOP拿数据（含日期），再距离+时间过滤
        _enrich_details(filtered_raw)
        
        pt = property_type or '商业'
        type_map = {'residential': '住宅', 'commercial': '商业', '住宅': '住宅', '商业': '商业'}
        mapped_type = type_map.get(pt, pt)
        dist_filtered = _filter_by_distance_time(filtered_raw, address, mapped_type)
        
        return _format_and_sort(dist_filtered, address)
    except Exception as e:
        print(f"搜索异常: {e}")
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'top3': [], 'all_cases': [],
            'self_auction_count': 0, 'total_count': 0,
            'error': '搜索服务暂时不可用，请稍后重试',
        }


@app.route('/api/search', methods=['POST'])
def search():
    """前端调用的主搜索接口"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'top3': [],
                'all_cases': [],
                'self_auction_count': 0,
                'total_count': 0,
                'error': '请求体为空或不是有效的JSON',
            }), 400

        address = data.get('address', '')
        if not address:
            return jsonify({
                'status': 'error',
                'top3': [],
                'all_cases': [],
                'self_auction_count': 0,
                'total_count': 0,
                'error': '地址不能为空',
            }), 400

        property_type = data.get('asset_type', data.get('propertyType', 'commercial'))
        area = data.get('building_area', data.get('area'))

        print(f"[API] 搜索请求: address={address}, type={property_type}, area={area}")
        result = run_search(address, property_type, area)

        if result.get('status') != 'success':
            print(f"[API] 搜索失败: {result.get('error', '未知错误')}")

        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[API] 未处理异常: {e}")
        return jsonify({
            'status': 'error',
            'top3': [],
            'all_cases': [],
            'self_auction_count': 0,
            'total_count': 0,
            'error': '服务器内部错误，请联系管理员',
        }), 500


# 保留旧接口兼容性
@app.route('/api/valuate', methods=['POST'])
def valuate():
    """旧版接口（兼容）"""
    data = request.get_json()
    address = data.get('address', '')
    property_type = data.get('propertyType', 'commercial')
    area = data.get('area')
    result = run_search(address, property_type, area)
    # 转换为旧格式
    return jsonify({
        'success': result['status'] == 'success',
        'message': result.get('error', '搜索成功') if result['status'] != 'success' else '搜索成功',
        'data': result.get('all_cases', []),
        'total': result.get('total_count', 0),
    })


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})

@app.route('/api/debug', methods=['GET'])
def debug():
    """调试接口 - 测试各模块是否可用"""
    deps = {}
    try:
        import playwright
        deps['playwright'] = 'ok'
        from playwright.sync_api import sync_playwright
        deps['playwright_api'] = 'ok'
    except Exception as e:
        deps['playwright'] = str(e)
    try:
        import playwright_searcher
        deps['playwright_searcher'] = 'ok'
    except Exception as e:
        deps['playwright_searcher'] = str(e)
    return jsonify(deps)


@app.route('/', methods=['GET'])
def index():
    return "不良资产估值参考案例搜索服务 - API运行中"


@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        raw = request.get_data(as_text=True)
        if not raw:
            return jsonify({'success': False, 'message': '请求体为空'}), 400
        data = json.loads(raw)
        cases = data.get('cases', data.get('all_cases', []))
        if not cases:
            return jsonify({'success': False, 'message': '没有可导出的数据'}), 400

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = '估值案例'

        headers = ['参照物位置', '土地面积 (m2)', '建筑面积 (m2)', '市场价值(万元)',
                   '建筑单价(元/m2)', '数据来源', '备注', '价格类型']
        col_widths = [60, 15, 15, 15, 18, 50, 80, 15]

        hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        hfont = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
        cfont = Font(name='Microsoft YaHei', size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        calign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        lalign = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hfill; cell.font = hfont; cell.alignment = calign; cell.border = border

        for ri, case in enumerate(cases, 2):
            vals = [
                case.get('referenceLocation') or case.get('参照物位置') or '',
                case.get('landArea') or case.get('土地面积') or '不适用',
                case.get('buildingArea') or case.get('建筑面积') or 0,
                case.get('marketValue') or case.get('市场价值') or 0,
                case.get('unitPrice') or case.get('建筑单价') or 0,
                case.get('link') or case.get('source') or case.get('数据来源') or '',
                case.get('remark') or case.get('备注') or '',
                case.get('priceType') or case.get('价格类型') or '普通司法拍卖',
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = cfont; cell.border = border
                cell.alignment = lalign if ci in (1, 6, 7) else calign
                if ci in (3, 4, 5):
                    try:
                        fv = float(v) if v not in ('-', '不适用', '', None) else 0
                        cell.value = fv
                    except: pass
                    cell.number_format = '#,##0.00'

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'

        import io
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=npl_valuation_cases.xlsx'})
    except Exception as e:
        print(f'导出异常: {e}')
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
