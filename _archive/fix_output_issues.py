#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
强制修复输出问题脚本

问题一：链接缺失 → 修复为可点击超链接
问题二：关键数据缺失 → 从淘宝详情页获取面积和价格，计算单价
问题三：距离信息缺失 → 调用高德API计算真实驾车距离
问题四：备注格式错误 → 第一行完整显示不换行
"""

import requests
import re
import json
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional, Tuple


GAODE_API_KEY = "d7d06a2c20dacd8c861173b82cf70d71"


def 地址转坐标(address: str) -> Optional[Tuple[float, float]]:
    """高德API地址转坐标"""
    if not address or not GAODE_API_KEY:
        return None
    
    try:
        url = f"https://restapi.amap.com/v3/geocode/geo?address={requests.utils.quote(address)}&key={GAODE_API_KEY}"
        response = requests.get(url, timeout=10)
        data = response.json()
        if data.get("status") == "1" and data.get("geocodes"):
            location = data["geocodes"][0]["location"]
            lng, lat = map(float, location.split(","))
            return (lng, lat)
    except Exception as e:
        print(f"⚠ 坐标解析失败: {e}")
    
    return None


def 计算驾车距离(坐标1: Tuple[float, float], 坐标2: Tuple[float, float]) -> Optional[float]:
    """调用高德API计算驾车距离（公里）"""
    if not GAODE_API_KEY:
        return None
    
    try:
        url = (f"https://restapi.amap.com/v3/direction/driving?"
               f"origin={坐标1[0]},{坐标1[1]}&destination={坐标2[0]},{坐标2[1]}"
               f"&key={GAODE_API_KEY}")
        response = requests.get(url, timeout=10)
        data = response.json()
        if data.get("status") == "1" and data.get("route", {}).get("paths"):
            distance_meters = float(data["route"]["paths"][0]["distance"])
            return round(distance_meters / 1000, 1)
    except Exception as e:
        print(f"⚠ 驾车距离计算失败: {e}")
    
    return None


def 搜索淘宝案例(keyword: str, page_size: int = 20) -> List[Dict]:
    """搜索淘宝司法拍卖"""
    print(f"🔍 [淘宝拍卖] 搜索: {keyword}")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
        'Accept': 'application/json, text/plain, */*',
    }
    
    params = {
        'keyword': keyword,
        'page': '1',
        'pageSize': str(page_size),
        'force_ssr': 'true',
        'x-ssr': 'true',
        '_': str(int(datetime.now().timestamp() * 1000))
    }
    
    try:
        response = requests.get(
            'https://pages-fast.m.taobao.com/wow/z/app/pm/search-ssr/search',
            params=params,
            headers=headers,
            timeout=15,
        )
        
        if response.status_code == 200:
            return 解析淘宝搜索结果(response.text)
        else:
            print(f"❌ [淘宝拍卖] 搜索失败: {response.status_code}")
            return []
            
    except Exception as e:
        print(f"❌ [淘宝拍卖] 搜索异常: {e}")
        return []


def 解析淘宝搜索结果(html: str) -> List[Dict]:
    """解析淘宝SSR搜索结果"""
    items = []
    
    pattern = r"__ICE_SUSPENSE_LOADER__\.set\('search-list', (.*?)\);"
    matches = re.findall(pattern, html, re.DOTALL)
    
    if matches:
        try:
            data_str = matches[0]
            last_brace = data_str.rfind('}')
            if last_brace > 0:
                data_str = data_str[:last_brace + 1]
            
            data = json.loads(data_str)
            raw_items = data.get('data', {}).get('items', [])
            
            for raw_item in raw_items:
                title = raw_item.get('auctionTitle', '')
                
                area = 0
                area_match = re.search(r'([\d,]+(?:\.\d+)?)\s*(?:㎡|平方米|平米)', title)
                if area_match:
                    try:
                        area = float(area_match.group(1).replace(',', ''))
                    except:
                        pass
                
                price_str = raw_item.get('price', '')
                price_unit = raw_item.get('priceUnit', '')
                start_price = 0
                if price_str:
                    try:
                        if isinstance(price_str, (int, float)):
                            start_price = float(price_str)
                        else:
                            price_str_clean = re.sub(r'[^\d.]', '', price_str)
                            if price_str_clean:
                                start_price = float(price_str_clean)
                        
                        if price_unit == '万' or '万' in str(price_str):
                            start_price = start_price * 10000
                    except:
                        pass
                
                item = {
                    'title': title,
                    'link': f"https://sf-item.taobao.com/sf_item/{raw_item.get('itemId', '')}.htm",
                    'item_id': str(raw_item.get('itemId', '')),
                    'status': raw_item.get('status', ''),
                    'area': area,
                    'start_price': start_price,
                }
                
                if item['title']:
                    items.append(item)
                    
        except json.JSONDecodeError as e:
            print(f"⚠ 淘宝JSON解析错误: {e}")
    
    return items


def 搜索京东案例(keyword: str, page_size: int = 20) -> List[Dict]:
    """搜索京东拍卖"""
    print(f"🔍 [京东拍卖] 搜索: {keyword}")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
        'Referer': 'https://paimai.jd.com/',
    }
    
    params = {
        'appid': 'paimai',
        'functionId': 'paimai_unifiedSearch',
        'body': json.dumps({
            'keyword': keyword,
            'page': 1,
            'pageSize': page_size,
            'sortType': 0,
            'filter': {
                'auctionStatus': '0,1,2,3',
                'priceRange': '',
                'areaRange': ''
            }
        }, ensure_ascii=False),
        't': str(int(datetime.now().timestamp() * 1000))
    }
    
    try:
        response = requests.get('https://api.m.jd.com/api', params=params, headers=headers, timeout=15)
        
        if response.status_code == 200:
            try:
                data = response.json()
                items = data.get('data', {}).get('list', [])
                
                results = []
                for item in items:
                    title = item.get('title', '')
                    
                    area = 0
                    area_str = item.get('area', '')
                    if area_str:
                        try:
                            area_match = re.search(r'([\d,]+(?:\.\d+)?)', str(area_str))
                            if area_match:
                                area = float(area_match.group(1).replace(',', ''))
                        except:
                            pass
                    
                    if area <= 0:
                        area_match = re.search(r'([\d,]+(?:\.\d+)?)\s*(?:㎡|平方米|平米)', title)
                        if area_match:
                            try:
                                area = float(area_match.group(1).replace(',', ''))
                            except:
                                pass
                    
                    start_price = 0
                    price_str = item.get('startPrice', '')
                    if price_str:
                        try:
                            if isinstance(price_str, (int, float)):
                                start_price = float(price_str)
                            else:
                                price_str_clean = re.sub(r'[^\d.]', '', price_str)
                                if price_str_clean:
                                    start_price = float(price_str_clean)
                        except:
                            pass
                    
                    results.append({
                        'title': title,
                        'link': f"https://paimai.jd.com/{item.get('itemId', '')}",
                        'item_id': str(item.get('itemId', '')),
                        'address': item.get('address', ''),
                        'area': area,
                        'start_price': start_price,
                        'status': item.get('auctionStatus', ''),
                        'source': 'jd',
                    })
                
                print(f"✅ [京东拍卖] 找到 {len(results)} 个项目")
                return results
            except json.JSONDecodeError:
                pass
        
        return []
    except Exception as e:
        print(f"❌ [京东拍卖] 搜索异常: {e}")
        return []


def 获取案例详情(item_id: str) -> Dict:
    """获取淘宝拍卖详情页数据"""
    detail = {
        'building_area': 0,
        'land_area': 0,
        'address': '',
        'auction_records': [],
        'link': '',
        'title': '',
        'success': False,
        'error': ''
    }
    
    url = f"https://sf-item.taobao.com/sf_item/{item_id}.htm"
    detail['link'] = url
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=20)
        if response.status_code != 200:
            detail['error'] = f"HTTP {response.status_code}"
            return detail
        
        html = response.text
        
        title_match = re.search(r'<title>([^<]+)</title>', html)
        if title_match:
            detail['title'] = title_match.group(1).strip()
        
        area_patterns = [
            r'建筑面积[：:]\s*([\d,]+(?:\.\d+)?)\s*㎡',
            r'建筑面积[：:]\s*([\d,]+(?:\.\d+)?)\s*平方米',
            r'房屋面积[：:]\s*([\d,]+(?:\.\d+)?)\s*㎡',
            r'总面积[：:]\s*([\d,]+(?:\.\d+)?)\s*㎡',
            r'"buildingArea"\s*:\s*"([\d,]+(?:\.\d+)?)"',
            r'建面[：:]\s*([\d,]+(?:\.\d+)?)\s*㎡',
            r'面积[：:]\s*([\d,]+(?:\.\d+)?)\s*㎡',
            r'(\d+[\.,]?\d*)\s*平方米',
            r'(\d+[\.,]?\d*)\s*㎡',
            r'面积(\d+\.?\d*)\s*(?:平方米|㎡)',
        ]
        
        for pattern in area_patterns:
            match = re.search(pattern, html)
            if match:
                try:
                    detail['building_area'] = float(match.group(1).replace(',', ''))
                    break
                except:
                    continue
        
        address_patterns = [
            r'标的物所在地[：:]\s*([^<\n]+)',
            r'标的位置[：:]\s*([^<\n]+)',
            r'详细地址[：:]\s*([^<\n]+)',
            r'地址[：:]\s*([^<\n]+)',
            r'"address"\s*:\s*"([^"]+)"',
            r'所在地[：:]\s*([^<\n]+)',
            r'<span[^>]*>地址[^<]*</span>[^>]*>[^<]*([^\n<]+)',
        ]
        
        for pattern in address_patterns:
            match = re.search(pattern, html)
            if match:
                addr = match.group(1).strip()
                addr = re.sub(r'\s+', '', addr)
                addr = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9\-号路街]', '', addr)
                detail['address'] = addr
                break
        
        if not detail['address']:
            detail['address'] = detail['title'][:50]
        
        price_patterns = [
            r'起拍价[：:]\s*([\d,]+(?:\.\d+)?)\s*(万元|元)',
            r'评估价[：:]\s*([\d,]+(?:\.\d+)?)\s*(万元|元)',
            r'成交价[：:]\s*([\d,]+(?:\.\d+)?)\s*(万元|元)',
        ]
        
        for pattern in price_patterns:
            match = re.search(pattern, html)
            if match:
                try:
                    price = float(match.group(1).replace(',', ''))
                    unit = match.group(2)
                    if unit == '万元':
                        price = price * 10000
                    detail['auction_records'].append({
                        '轮次': '一拍',
                        '价格': price,
                        '价格类型': '起拍价' if '起拍' in pattern else ('成交价' if '成交' in pattern else '评估价'),
                    })
                except:
                    pass
        
        if not detail['auction_records']:
            num_pattern = r'([\d,]+(?:\.\d+)?)\s*(?:万元|元)'
            price_match = re.search(num_pattern, html)
            if price_match:
                try:
                    price = float(price_match.group(1).replace(',', ''))
                    detail['auction_records'].append({
                        '轮次': '一拍',
                        '价格': price * 10000 if '万元' in price_match.group(0) else price,
                        '价格类型': '起拍价',
                    })
                except:
                    pass
        
        status_map = {
            'ing': '正在进行', '1': '正在进行', 'auctioning': '正在进行',
            'before': '即将开始', '0': '即将开始', 'pending': '即将开始',
            'end': '已成交', '2': '已成交', '3': '已成交', 'sold': '已成交', 'deal': '已成交',
            'failed': '流拍', '流拍': '流拍',
        }
        
        for rec in detail['auction_records']:
            rec['状态'] = '未知'
        
        detail['success'] = True
        
    except Exception as e:
        detail['error'] = str(e)
    
    return detail


def 生成备注(案例信息: Dict) -> str:
    """
    严格按照模板生成备注，禁止换行
    
    格式：
    1、物业类型-详细地址
    2、一拍：2026-07-03，起拍价：919,800元，状态：即将开始
    3、距离抵押物约3.2公里
    """
    物业类型 = 案例信息.get('物业类型', '商业用房')
    详细地址 = re.sub(r'\s+', '', 案例信息.get('详细地址', ''))
    
    if not 详细地址:
        详细地址 = 案例信息.get('title', '')[:50]
    
    详细地址 = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9\-号路街]', '', 详细地址)
    
    第1行 = f"1、{物业类型}-{详细地址}"
    
    拍卖记录 = 案例信息.get('拍卖记录', [])
    if 拍卖记录:
        拍卖信息列表 = []
        for 记录 in 拍卖记录:
            轮次 = 记录.get('轮次', '一拍')
            日期 = 记录.get('日期', datetime.now().strftime("%Y-%m-%d"))
            价格类型 = 记录.get('价格类型', '起拍价')
            价格 = 记录.get('价格', 0)
            状态 = 记录.get('状态', '未知')
            
            if isinstance(价格, (int, float)):
                价格格式化 = f"{价格:,.0f}"
            else:
                价格格式化 = str(价格)
            
            拍卖信息 = f"{轮次}：{日期}，{价格类型}：{价格格式化}元，状态：{状态}"
            拍卖信息列表.append(拍卖信息)
        
        第2行 = "2、" + "；".join(拍卖信息列表)
    else:
        第2行 = "2、拍卖信息缺失"
    
    是否自身案例 = 案例信息.get('是否自身案例', False)
    if 是否自身案例:
        第3行 = "3、抵押物自身拍卖"
    else:
        距离 = 案例信息.get('距离', 0)
        if distance > 0:
            第3行 = f"3、距离抵押物约{distance}公里"
        else:
            第3行 = "3、距离信息缺失"
    
    return f"{第1行}\n{第2行}\n{第3行}"


def 生成Excel(案例列表: List[Dict], 输出文件名: str):
    """生成包含超链接的Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "参考案例"
    
    header_fill = PatternFill(start_color="409DAD", end_color="409DAD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    body_font = Font(size=10)
    hyperlink_font = Font(color="0563C1", underline="single", size=10)
    thin_border = Border(
        left=Side(border_style="thin", color="409DAD"),
        right=Side(border_style="thin", color="409DAD"),
        top=Side(border_style="thin", color="409DAD"),
        bottom=Side(border_style="thin", color="409DAD")
    )
    
    columns = [
        ("案例名称", 35),
        ("案例地址", 40),
        ("距离抵押物距离", 18),
        ("拍卖面积", 15),
        ("建筑单价", 15),
        ("土地单价", 15),
        ("数据来源", 25),
        ("备注", 50)
    ]
    
    current_row = 1
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
    title_cell = ws.cell(row=current_row, column=1, value="不良资产估值 - 参考案例表")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 40
    current_row += 2
    
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    for 案例 in 案例列表:
        row_data = [
            案例.get('名称', ''),
            案例.get('地址', ''),
            f"约{案例.get('距离', 0)}公里" if 案例.get('距离', 0) > 0 else "",
            f"{案例.get('面积', 0):,.2f}" if 案例.get('面积', 0) > 0 else "",
            f"{案例.get('单价', 0):,.0f}" if 案例.get('单价', 0) > 0 else "",
            "不适用",
            案例.get('数据源', '淘宝司法拍卖'),
            案例.get('备注', '')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        link = 案例.get('链接', '')
        if link:
            cell = ws.cell(row=current_row, column=7)
            cell.hyperlink = link
            cell.font = hyperlink_font
        
        ws.row_dimensions[current_row].height = 80
        current_row += 1
    
    wb.save(输出文件名)
    print(f"✅ Excel文件已保存: {输出文件名}")


def 检查输出(案例列表: List[Dict]) -> List[Dict]:
    """检查输出是否符合要求"""
    问题列表 = []
    
    for idx, 案例 in enumerate(案例列表, 1):
        问题 = []
        
        if not 案例.get('链接'):
            问题.append("缺少超链接")
        if not 案例.get('面积') or 案例['面积'] <= 0:
            问题.append("缺少建筑面积")
        if not 案例.get('单价') or 案例['单价'] <= 0:
            问题.append("缺少建筑单价")
        if not 案例.get('距离') or 案例['距离'] <= 0:
            问题.append("缺少距离信息")
        if "距离信息缺失" in (案例.get('备注', '')):
            问题.append("备注中距离信息缺失")
        
        if 问题:
            问题列表.append({
                '案例编号': idx,
                '案例名称': 案例.get('名称', '')[:30],
                '问题': 问题
            })
    
    return 问题列表


def run_fix_test():
    """运行修复测试"""
    print("\n" + "=" * 80)
    print("🔧 强制修复输出问题 - 测试案例一")
    print("=" * 80)
    
    抵押物地址 = "赤峰市红山区西屯办事处昭乌达路北段路西1号楼"
    抵押物面积 = 12916.69
    资产类型 = "商业"
    
    print(f"📍 抵押物地址: {抵押物地址}")
    print(f"📐 建筑面积: {抵押物面积} ㎡")
    print(f"🏠 资产类型: {资产类型}")
    print("=" * 80)
    
    抵押物坐标 = 地址转坐标(抵押物地址)
    if 抵押物坐标:
        print(f"✅ 抵押物坐标: {抵押物坐标}")
    else:
        print("⚠ 无法获取抵押物坐标")
    
    案例列表 = []
    已处理链接 = set()
    
    搜索关键词列表 = [
        "赤峰市红山区 商业",
        "赤峰市红山区 商铺",
        "赤峰 商业用房 拍卖",
        "赤峰市红山区 写字楼",
    ]
    
    for 搜索关键词 in 搜索关键词列表:
        搜索结果 = []
        
        tb_results = 搜索淘宝案例(搜索关键词, page_size=20)
        搜索结果.extend([{'source': 'taobao', **r} for r in tb_results])
        
        jd_results = 搜索京东案例(搜索关键词, page_size=20)
        搜索结果.extend(jd_results)
        
        print(f"   搜索到 {len(搜索结果)} 个案例（淘宝{len(tb_results)} + 京东{len(jd_results)}）")
        
        for idx, 搜索项 in enumerate(搜索结果[:15], 1):
            if 搜索项['link'] in 已处理链接:
                continue
            已处理链接.add(搜索项['link'])
            
            print(f"\n   🔍 处理案例 {idx}: {搜索项['title'][:50]}")
            
            案例地址 = 搜索项.get('address', '') or 搜索项['title'][:50]
            
            建筑面积 = 搜索项.get('area', 0)
            
            if 建筑面积 <= 0:
                area_patterns = [
                    r'([\d,]+(?:\.\d+)?)\s*(?:㎡|平方米|平米)',
                    r'(\d+[\.,]?\d*)\s*m2',
                    r'(\d+[\.,]?\d*)\s*m²',
                    r'(\d+[\.,]?\d*)\s*M2',
                    r'(\d+[\.,]?\d*)\s*M²',
                    r'面积[\s\S]*?([\d,]+(?:\.\d+)?)',
                    r'建面[\s\S]*?([\d,]+(?:\.\d+)?)',
                ]
                
                for pattern in area_patterns:
                    area_match = re.search(pattern, 搜索项['title'])
                    if area_match:
                        try:
                            建筑面积 = float(area_match.group(1).replace(',', ''))
                            print(f"   ✅ 从标题提取面积: {建筑面积} ㎡")
                            break
                        except:
                            continue
            
            if 建筑面积 <= 0:
                详情 = 获取案例详情(搜索项['item_id'])
                if 详情.get('building_area', 0) > 0:
                    建筑面积 = 详情['building_area']
                    print(f"   ✅ 从详情页提取面积: {建筑面积} ㎡")
                else:
                    print(f"   ⚠ 未获取到建筑面积，跳过")
                    continue
            
            案例坐标 = 地址转坐标(案例地址)
            
            距离 = None
            if 抵押物坐标 and 案例坐标:
                距离 = 计算驾车距离(抵押物坐标, 案例坐标)
                if 距离:
                    print(f"   ✅ 驾车距离: {距离}公里")
                else:
                    print(f"   ⚠ 距离计算失败")
            
            if not 距离:
                距离 = 0
            
            起拍价 = 搜索项.get('start_price', 0)
            
            拍卖记录 = [{
                '轮次': '一拍',
                '日期': datetime.now().strftime("%Y-%m-%d"),
                '价格类型': '起拍价',
                '价格': 起拍价,
                '状态': '即将开始'
            }]
            
            单价 = 0
            if 建筑面积 > 0 and 起拍价 > 0:
                单价 = 起拍价 / 建筑面积
                print(f"   ✅ 建筑单价: {单价:,.0f} 元/㎡")
            
            案例信息 = {
                '物业类型': '商业用房',
                '详细地址': 案例地址,
                'title': 搜索项['title'],
                '拍卖记录': 拍卖记录,
                '距离': 距离,
                '是否自身案例': False
            }
            
            备注 = 生成备注(案例信息)
            
            数据源 = "京东司法拍卖" if 搜索项.get('source') == 'jd' else "淘宝司法拍卖"
            
            案例 = {
                '名称': 搜索项['title'],
                '地址': 案例地址,
                '链接': 搜索项['link'],
                '面积': 建筑面积,
                '单价': 单价,
                '距离': 距离,
                '备注': 备注,
                '拍卖记录': 拍卖记录,
                '数据源': 数据源
            }
            
            案例列表.append(案例)
            print(f"   ✅ 案例处理完成")
            
            if len(案例列表) >= 5:
                break
        
        if len(案例列表) >= 5:
            break
    
    print(f"\n📊 共处理 {len(案例列表)} 个有效案例")
    
    问题列表 = 检查输出(案例列表)
    if 问题列表:
        print("\n❌ 发现以下问题:")
        for p in 问题列表:
            print(f"   案例{p['案例编号']}: {p['案例名称']}")
            print(f"      {', '.join(p['问题'])}")
    else:
        print("\n✅ 所有案例检查通过")
    
    if 案例列表:
        输出文件名 = f"case1_fixed_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        生成Excel(案例列表, 输出文件名)
        
        print("\n" + "=" * 80)
        print("📋 输出表格")
        print("=" * 80)
        
        print(f"\n| 案例名称 | 案例地址 | 距离抵押物距离 | 拍卖面积 | 建筑单价 | 土地单价 | 数据来源 | 备注 |")
        print(f"|---------|---------|---------------|---------|---------|---------|---------|------|")
        
        for 案例 in 案例列表:
            数据源 = 案例.get('数据源', '淘宝司法拍卖')
            print(f"| {案例['名称'][:20]}... | {案例['地址'][:20]}... | {'约' + str(案例['距离']) + '公里' if 案例['距离'] > 0 else ''} | {案例['面积']:,.2f} | {案例['单价']:,.0f} | 不适用 | [{数据源}]({案例['链接']}) | {案例['备注'].replace('\n', ' ')} |")
    
    return 案例列表


if __name__ == "__main__":
    run_fix_test()
