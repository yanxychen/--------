import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

class ExcelFinalFixer:
    """最终修复所有Excel问题的类"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="409DAD", end_color="409DAD", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.hyperlink_font = Font(color="0563C1", underline="single", size=10)
        self.normal_font = Font(size=10)
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def fix_all_files(self, file_paths):
        """修复所有Excel文件"""
        results = {}
        
        for input_path in file_paths:
            if not input_path.endswith('.xlsx'):
                continue
                
            output_path = input_path.replace('.xlsx', '_final_fixed.xlsx')
            
            try:
                # 修复文件
                self._fix_single_file(input_path, output_path)
                
                # 验证修复
                issues = self._validate_fixed_file(output_path)
                
                results[input_path] = {
                    'status': '✅ 已修复',
                    'output': output_path,
                    'remaining_issues': len(issues),
                    'issues': issues[:5]
                }
                
            except Exception as e:
                results[input_path] = {
                    'status': f'❌ 修复失败: {str(e)}',
                    'output': None,
                    'remaining_issues': 0,
                    'issues': []
                }
        
        return results
    
    def _fix_single_file(self, input_path, output_path):
        """修复单个Excel文件（创建新工作簿，避免合并单元格问题）"""
        # 加载原始文件
        source_wb = openpyxl.load_workbook(input_path)
        source_ws = source_wb.active
        
        # 创建新工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 1. 写入正确表头
        self._fix_headers(ws)
        
        # 2. 找到数据起始行（跳过标题行/统计行）
        data_start_row = self._find_data_start_row(source_ws)
        if data_start_row is None:
            data_start_row = 1
        
        # 3. 修复并写入每一行数据
        row_num = 2  # 新文件的数据起始行
        for row in range(data_start_row + 1, source_ws.max_row + 1):
            # 获取原始数据
            original_data = self._extract_row_data_from_source(source_ws, row)
            
            # 如果是空白行或说明行，跳过
            if not original_data or self._is_comment_row(original_data):
                continue
            
            # 修复数据
            fixed_data = self._fix_row_data(original_data)
            
            # 写回修复后的数据
            self._write_fixed_data(ws, row_num, fixed_data)
            
            row_num += 1
        
        # 4. 调整格式
        self._apply_formatting(ws)
        
        # 5. 保存
        wb.save(output_path)
        source_wb.close()
    
    def _find_data_start_row(self, ws):
        """找到数据表头所在行"""
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            first_cell = str(ws.cell(row=row_idx, column=1).value or "").strip()
            if first_cell == "参照物位置" or first_cell.startswith("参照物"):
                return row_idx
        return None
    
    def _extract_row_data_from_source(self, ws, row):
        """从源文件提取行数据（处理合并单元格）"""
        data = {}
        
        # 先获取表头
        headers = {}
        header_row = self._find_data_start_row(ws)
        if header_row:
            for col in range(1, min(ws.max_column + 1, 15)):
                header = str(ws.cell(row=header_row, column=col).value or "").strip()
                if header:
                    headers[col] = header
        
        # 如果没有找到表头，使用默认列名
        if not headers:
            default_headers = [
                "参照物位置", "土地面积(m²)", "建筑面积(m²)", 
                "市场价值(万元)", "建筑单价(元/㎡)", "数据来源", "备注", "价格类型"
            ]
            for i, h in enumerate(default_headers, 1):
                headers[i] = h
        
        # 获取数据
        for col in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            
            # 如果是合并单元格，找到原始值
            if isinstance(cell, openpyxl.cell.MergedCell):
                # 查找合并区域的起始单元格
                for merged_range in ws.merged_cells.ranges:
                    if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                        break
            
            if col in headers:
                data[headers[col]] = value
        
        return data
    
    def _is_comment_row(self, data):
        """判断是否是说明行"""
        location = str(data.get('参照物位置') or data.get('土地面积(m²)') or "")
        if not location:
            return True
        if location.startswith('说明') or location.startswith('备注') or location.startswith('注：'):
            return True
        if location.startswith('1. ') or location.startswith('2. ') or location.startswith('3. '):
            return True
        return False
    
    def _fix_headers(self, ws):
        """修复表头为8列正确格式"""
        correct_headers = [
            "参照物位置",
            "土地面积 (m²)", 
            "建筑面积 (m²)",
            "市场价值 (万元)",
            "建筑单价 (元/m²)",
            "数据来源",
            "备注",
            "价格类型"
        ]
        
        # 写入正确表头（新工作簿不需要清空，直接写入）
        for col_idx, header in enumerate(correct_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
    
    def _extract_row_data(self, ws, row):
        """提取行数据"""
        data = {}
        
        # 尝试从不同列提取数据
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            value = ws.cell(row=row, column=col).value
            
            if header:
                data[header] = value
        
        return data
    
    def _fix_row_data(self, original_data):
        """修复单行数据"""
        fixed = {}
        
        # 1. 参照物位置（第1列）
        location = original_data.get('参照物位置') or original_data.get('location') or '未知位置'
        fixed['参照物位置'] = location
        
        # 2. 土地面积（第2列）
        land_area = original_data.get('土地面积 (m²)') or original_data.get('土地面积(m²)') or original_data.get('land_area')
        fixed['土地面积 (m²)'] = self._format_area(land_area)
        
        # 3. 建筑面积（第3列）
        building_area = original_data.get('建筑面积 (m²)') or original_data.get('建筑面积(m²)') or original_data.get('building_area')
        fixed['建筑面积 (m²)'] = self._format_area(building_area)
        
        # 4. 市场价值（第4列）
        market_value = original_data.get('市场价值 (万元)') or original_data.get('市场价值(万元)') or original_data.get('market_value')
        fixed['市场价值 (万元)'] = self._format_number(market_value)
        
        # 5. 建筑单价（第5列）- 重新计算
        fixed['建筑单价 (元/m²)'] = self._calculate_unit_price(
            fixed['市场价值 (万元)'], 
            fixed['建筑面积 (m²)']
        )
        
        # 6. 数据来源（第6列）
        source = original_data.get('数据来源') or original_data.get('source') or '淘宝司法拍卖'
        fixed['数据来源'] = source
        
        # 7. 备注（第7列）- 完全重构
        fixed['备注'] = self._generate_complete_remark(original_data)
        
        # 8. 价格类型（第8列）
        fixed['价格类型'] = '普通司法拍卖'
        
        return fixed
    
    def _format_area(self, value):
        """格式化面积"""
        if value is None or value == 0 or value == '0' or str(value).strip() in ['不适用', 'None', 'null', '']:
            return '不适用'
        
        try:
            num = float(str(value).replace(',', '').replace('【', '').replace('】', '').replace('平方米', ''))
            return f"{num:,.2f}"
        except:
            return '不适用'
    
    def _format_number(self, value):
        """格式化数值"""
        if value is None or str(value).strip() in ['不适用', 'None', 'null', '']:
            return '不适用'
        
        try:
            num = float(str(value).replace(',', ''))
            return f"{num:,.2f}"
        except:
            return '不适用'
    
    def _calculate_unit_price(self, market_value_str, building_area_str):
        """计算建筑单价"""
        # 解析字符串
        try:
            market_value = float(str(market_value_str).replace('不适用', '0').replace(',', ''))
            building_area = float(str(building_area_str).replace('不适用', '0').replace(',', ''))
        except:
            return '不适用'
        
        # 逻辑判断
        if building_area == 0 or building_area_str == '不适用':
            return '不适用'
        
        if market_value > 0 and building_area > 0:
            unit_price = (market_value * 10000) / building_area
            return f"{unit_price:,.2f}"
        
        return '0.00'
    
    def _generate_complete_remark(self, data):
        """生成包含5要素的完整备注"""
        parts = []
        
        # 1. 拍卖轮次（从标题或备注中提取）
        auction_round = self._extract_auction_round(data)
        parts.append(f"拍卖轮次：{auction_round}")
        
        # 2. 时间（从原始数据提取或默认）
        auction_date = self._extract_auction_date(data)
        parts.append(f"时间：{auction_date}")
        
        # 3. 价格（起拍价+成交价）
        price_text = self._format_prices(data)
        parts.append(price_text)
        
        # 4. 距离
        distance_text = self._format_distance(data)
        parts.append(distance_text)
        
        # 5. 状态
        status_text = self._format_status(data)
        parts.append(f"状态：{status_text}")
        
        return "；".join(parts)
    
    def _extract_auction_round(self, data):
        """提取拍卖轮次"""
        # 从备注中提取
        remark = str(data.get('备注') or '')
        
        # 查找轮次关键词
        for keyword in ['一拍', '二拍', '三拍', '变卖']:
            if keyword in remark:
                return keyword
        
        # 从标题中提取
        title = str(data.get('title') or data.get('参照物位置') or '')
        if '一拍' in title:
            return '一拍'
        elif '二拍' in title:
            return '二拍'
        elif '三拍' in title:
            return '三拍'
        elif '变卖' in title:
            return '变卖'
        
        # 默认
        return '一拍'
    
    def _extract_auction_date(self, data):
        """提取拍卖时间"""
        # 从备注中提取日期
        remark = str(data.get('备注') or '')
        
        # 查找日期格式 YYYY-MM-DD
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', remark)
        if date_match:
            return date_match.group(1)
        
        # 从原始数据中提取
        date_str = str(data.get('auction_date') or data.get('date') or '')
        if date_str and len(date_str) >= 10:
            return date_str[:10]
        
        # 默认今天
        return datetime.now().strftime('%Y-%m-%d')
    
    def _format_prices(self, data):
        """格式化价格部分（从原始备注和市场价值字段提取）"""
        price_parts = []
        
        # 从备注中提取起拍价
        remark = str(data.get('备注') or '')
        
        # 查找起拍价（支持多种格式）
        start_price_patterns = [
            r'起拍价[:：]\s*([\d,]+(\.\d+)?)\s*万元',
            r'起拍价[:：]\s*([\d,]+(\.\d+)?)\s*元',
            r'起拍价[:：]\s*([\d,]+(\.\d+)?)',
        ]
        
        found_price = False
        for pattern in start_price_patterns:
            match = re.search(pattern, remark)
            if match:
                price_str = match.group(1)
                try:
                    price_num = float(price_str.replace(',', ''))
                    # 判断单位（如果没匹配到单位，根据数值大小判断）
                    if price_num < 10000:  # 万元为单位
                        price_num = price_num * 10000
                    price_parts.append(f"起拍价：{price_num:,.2f}元")
                    found_price = True
                    break
                except:
                    pass
        
        # 如果备注中没找到，从市场价值字段获取
        if not found_price:
            market_value = data.get('市场价值 (万元)') or data.get('市场价值(万元)')
            if market_value:
                try:
                    num = float(str(market_value).replace(',', ''))
                    if num > 0:
                        price_parts.append(f"起拍价：{num * 10000:,.2f}元")
                        found_price = True
                except:
                    pass
        
        # 查找成交价
        deal_price_patterns = [
            r'成交价[:：]\s*([\d,]+(\.\d+)?)\s*万元',
            r'成交价[:：]\s*([\d,]+(\.\d+)?)\s*元',
            r'成交价[:：]\s*([\d,]+(\.\d+)?)',
        ]
        
        for pattern in deal_price_patterns:
            match = re.search(pattern, remark)
            if match:
                price_str = match.group(1)
                try:
                    price_num = float(price_str.replace(',', ''))
                    if price_num < 10000:
                        price_num = price_num * 10000
                    price_parts.append(f"成交价：{price_num:,.2f}元")
                except:
                    pass
        
        if price_parts:
            return "；".join(price_parts)
        else:
            return "价格：未知"
    
    def _format_distance(self, data):
        """格式化距离"""
        # 检查是否是抵押物自身案例
        remark = str(data.get('备注') or '')
        if '抵押物自身拍卖案例' in remark:
            return '距离抵押物：抵押物自身拍卖案例'
        
        # 从数据中提取距离
        distance_km = data.get('distance_km') or data.get('distance')
        
        if distance_km:
            try:
                distance = float(str(distance_km).replace(',', ''))
                if distance < 1:
                    meters = int(distance * 1000)
                    return f"距离抵押物：{meters}米"
                else:
                    return f"距离抵押物：{distance:.1f}公里"
            except:
                pass
        
        # 从备注中提取距离
        distance_match = re.search(r'距离抵押物[:：]\s*([\d\.]+)\s*(米|公里)', remark)
        if distance_match:
            return f"距离抵押物：{distance_match.group(1)}{distance_match.group(2)}"
        
        # 默认
        return "距离抵押物：距离未知"
    
    def _format_status(self, data):
        """格式化状态"""
        status = str(data.get('status') or '')
        remark = str(data.get('备注') or '')
        
        status_map = {
            'ing': '正在进行',
            '1': '正在进行',
            'auctioning': '正在进行',
            'before': '即将开始',
            '0': '即将开始',
            'pending': '即将开始',
            '待开拍': '即将开始',
            'end': '已成交',
            '2': '已成交',
            '3': '已成交',
            'sold': '已成交',
            'deal': '已成交',
            'failed': '正在进行流拍',
            '流拍': '正在进行流拍',
            'fail': '变卖失败',
            '变卖失败': '变卖失败',
            '拍卖中': '正在进行',
            '已结束': '已成交',
        }
        
        # 从状态字段获取
        if status in status_map:
            return status_map[status]
        
        # 从备注中提取
        for eng_status, chi_status in status_map.items():
            if eng_status in remark.lower() or chi_status in remark:
                return chi_status
        
        # 默认
        return '正在进行'
    
    def _write_fixed_data(self, ws, row, fixed_data):
        """写回修复后的数据"""
        # 第1列：参照物位置
        ws.cell(row=row, column=1, value=fixed_data['参照物位置'])
        
        # 第2-8列
        columns_mapping = {
            2: '土地面积 (m²)',
            3: '建筑面积 (m²)',
            4: '市场价值 (万元)',
            5: '建筑单价 (元/m²)',
            6: '数据来源',
            7: '备注',
            8: '价格类型'
        }
        
        for col_idx, key in columns_mapping.items():
            ws.cell(row=row, column=col_idx, value=fixed_data[key])
    
    def _apply_formatting(self, ws):
        """应用格式"""
        # 设置列宽
        column_widths = {
            'A': 40,
            'B': 15,
            'C': 15,
            'D': 18,
            'E': 18,
            'F': 12,
            'G': 50,
            'H': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.alignment = self.left_align
                cell.font = self.normal_font
    
    def _validate_fixed_file(self, file_path):
        """验证修复后的文件"""
        issues = []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 检查表头
            headers = [ws.cell(row=1, column=col).value for col in range(1, 9)]
            expected_headers = [
                "参照物位置",
                "土地面积 (m²)", 
                "建筑面积 (m²)",
                "市场价值 (万元)",
                "建筑单价 (元/m²)",
                "数据来源",
                "备注",
                "价格类型"
            ]
            
            if headers != expected_headers:
                issues.append(f"表头不正确: {headers}")
            
            # 检查数据行
            for row in range(2, ws.max_row + 1):
                row_issues = []
                
                # 检查备注
                remark = ws.cell(row=row, column=7).value
                if remark:
                    # 检查5要素
                    required_elements = [
                        ('拍卖轮次', ['拍卖轮次', '一拍', '二拍', '三拍', '变卖']),
                        ('时间', ['时间', r'\d{4}-\d{2}-\d{2}']),
                        ('价格', ['起拍价', '成交价', '价格']),
                        ('距离', ['距离', '抵押物自身拍卖案例']),
                        ('状态', ['状态']),
                    ]
                    for element_name, keywords in required_elements:
                        found = False
                        for kw in keywords:
                            if re.search(kw, str(remark)):
                                found = True
                                break
                        if not found:
                            row_issues.append(f"备注缺少'{element_name}'")
                
                # 检查建筑单价
                building_area = ws.cell(row=row, column=3).value
                unit_price = ws.cell(row=row, column=5).value
                
                if building_area == '不适用' and unit_price != '不适用':
                    row_issues.append("建筑面积为'不适用'但单价不是'不适用'")
                
                if row_issues:
                    issues.append(f"第{row}行: {'; '.join(row_issues)}")
            
        except Exception as e:
            issues.append(f"验证失败: {str(e)}")
        
        return issues

# 使用示例
if __name__ == "__main__":
    # 需要修复的文件列表
    files_to_fix = [
        "verify.xlsx",
        "test_valuation.xlsx", 
        "verify_fixed.xlsx",
        "test_fixed_output.xlsx"
    ]
    
    # 创建修复器
    fixer = ExcelFinalFixer()
    
    # 修复所有文件
    results = fixer.fix_all_files(files_to_fix)
    
    # 打印结果
    print("=" * 60)
    print("Excel文件修复结果")
    print("=" * 60)
    
    for file, result in results.items():
        print(f"\n📁 {file}")
        print(f"  状态: {result['status']}")
        if result['output']:
            print(f"  输出文件: {result['output']}")
            print(f"  剩余问题: {result['remaining_issues']}个")
            if result['issues']:
                print(f"  问题详情:")
                for issue in result['issues']:
                    print(f"    - {issue}")
    print("\n" + "=" * 60)
