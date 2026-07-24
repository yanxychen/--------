#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel渲染器 - 8列标准格式输出

列格式:
1. 参照物位置
2. 土地面积(m²)
3. 建筑面积(m²)
4. 市场价值(万元)
5. 建筑单价(元/㎡)
6. 数据来源
7. 备注
8. 价格类型
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional


class ExcelRenderer:
    """Excel渲染器"""
    
    def __init__(self):
        self.header_fill = PatternFill(
            start_color="409DAD",
            end_color="409DAD",
            fill_type="solid"
        )
        self.header_font = Font(
            name="微软雅黑",
            size=11,
            bold=True,
            color="FFFFFF"
        )
        self.body_font = Font(
            name="微软雅黑",
            size=10
        )
        self.border_color = "409DAD"
        self.thin_border = Border(
            left=Side(border_style="thin", color=self.border_color),
            right=Side(border_style="thin", color=self.border_color),
            top=Side(border_style="thin", color=self.border_color),
            bottom=Side(border_style="thin", color=self.border_color)
        )
        self.outer_border = Border(
            left=Side(border_style="medium", color=self.border_color),
            right=Side(border_style="medium", color=self.border_color),
            top=Side(border_style="medium", color=self.border_color),
            bottom=Side(border_style="medium", color=self.border_color)
        )
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        self.left_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
        
        self.columns = [
            ("参照物位置", 40),
            ("土地面积(m²)", 15),
            ("建筑面积(m²)", 15),
            ("市场价值(万元)", 15),
            ("建筑单价(元/㎡)", 18),
            ("数据来源", 50),
            ("备注", 40),
            ("价格类型", 15)
        ]
    
    def render_to_excel(self, cases: List[Dict], output_path: str, 
                       title: str = "不良资产估值 - 参考案例表",
                       statistics: Optional[Dict] = None):
        """
        渲染到Excel文件
        
        Args:
            cases: 案例列表（8列格式）
            output_path: 输出文件路径
            title: 表格标题
            statistics: 统计信息
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "参考案例"
        
        current_row = 1
        
        # 标题行
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=len(self.columns))
        title_cell = ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(name="微软雅黑", size=16, bold=True)
        title_cell.alignment = self.center_alignment
        title_cell.fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
        # 统计信息行（如果有）
        if statistics:
            stats_text = self._format_statistics(statistics)
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(self.columns))
            stats_cell = ws.cell(row=current_row, column=1, value=stats_text)
            stats_cell.font = Font(name="微软雅黑", size=10)
            stats_cell.alignment = self.left_alignment
            ws.row_dimensions[current_row].height = 30
            current_row += 1
        
        # 空行
        current_row += 1
        
        # 表头行
        header_row = current_row
        for col_idx, (col_name, width) in enumerate(self.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[header_row].height = 30
        current_row += 1
        
        # 数据行
        for case_idx, case in enumerate(cases, 1):
            for col_idx, (col_name, width) in enumerate(self.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=case.get(col_name, ""))
                cell.font = self.body_font
                cell.border = self.thin_border
                
                if col_idx in [1, 6, 7]:
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment
            
            # 为参照物位置列添加超链接
            link_url = case.get("参照物位置_链接", "")
            if link_url:
                cell = ws.cell(row=current_row, column=1)
                cell.hyperlink = link_url
                cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
            
            # 为数据来源列添加超链接
            source_url = case.get("数据来源_链接", "") or case.get("source_url", "") or case.get("link", "")
            if source_url:
                cell = ws.cell(row=current_row, column=6)
                cell.hyperlink = source_url
                cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
            
            ws.row_dimensions[current_row].height = 25
            current_row += 1
        
        # 底部空行
        current_row += 1
        
        # 说明行
        notes = [
            "说明：",
            "1. 以上案例来源于京东拍卖和淘宝司法拍卖，仅供参考",
            "2. 市场价值为拍卖当前价或起拍价，非最终成交价",
            "3. 建筑单价 = 市场价值 / 建筑面积",
            "4. 价格类型：普通司法拍卖"
        ]
        
        for note in notes:
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(self.columns))
            note_cell = ws.cell(row=current_row, column=1, value=note)
            note_cell.font = Font(name="微软雅黑", size=9, color="666666")
            note_cell.alignment = self.left_alignment
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        # 外边框
        max_row = current_row - 1
        for col_idx in range(1, len(self.columns) + 1):
            ws.cell(row=header_row, column=col_idx).border = Border(
                left=Side(border_style="thin", color=self.border_color),
                right=Side(border_style="thin", color=self.border_color),
                top=Side(border_style="medium", color=self.border_color),
                bottom=Side(border_style="thin", color=self.border_color)
            )
            ws.cell(row=header_row + len(cases), column=col_idx).border = Border(
                left=Side(border_style="thin", color=self.border_color),
                right=Side(border_style="thin", color=self.border_color),
                top=Side(border_style="thin", color=self.border_color),
                bottom=Side(border_style="medium", color=self.border_color)
            )
        
        for row_idx in range(header_row, header_row + len(cases) + 1):
            ws.cell(row=row_idx, column=1).border = Border(
                left=Side(border_style="medium", color=self.border_color),
                right=Side(border_style="thin", color=self.border_color),
                top=Side(border_style="thin", color=self.border_color),
                bottom=Side(border_style="thin", color=self.border_color)
            )
            ws.cell(row=row_idx, column=len(self.columns)).border = Border(
                left=Side(border_style="thin", color=self.border_color),
                right=Side(border_style="medium", color=self.border_color),
                top=Side(border_style="thin", color=self.border_color),
                bottom=Side(border_style="thin", color=self.border_color)
            )
        
        # 保存
        wb.save(output_path)
        print(f"✅ Excel文件已保存: {output_path}")
        print(f"   共 {len(cases)} 个案例，8列标准格式")
    
    def _format_statistics(self, statistics: Dict) -> str:
        """格式化统计信息"""
        parts = []
        
        if statistics.get("case_count"):
            parts.append(f"案例数量：{statistics['case_count']} 个")
        
        if statistics.get("reference_avg_price"):
            parts.append(f"参考均价：{statistics['reference_avg_price']} 元/㎡")
        
        if statistics.get("min_unit_price") and statistics.get("max_unit_price"):
            parts.append(f"单价区间：{statistics['min_unit_price']} ~ {statistics['max_unit_price']} 元/㎡")
        
        if statistics.get("avg_market_value"):
            parts.append(f"平均市值：{statistics['avg_market_value']} 万元")
        
        return "    ".join(parts)
    
    def create_valuation_report(self, cases: List[Dict], output_path: str,
                               asset_info: Optional[Dict] = None,
                               statistics: Optional[Dict] = None):
        """
        创建完整的估值报告
        
        Args:
            cases: 案例列表
            output_path: 输出文件路径
            asset_info: 标的资产信息
            statistics: 统计信息
        """
        wb = openpyxl.Workbook()
        
        # Sheet1: 参考案例表
        ws1 = wb.active
        ws1.title = "参考案例"
        self._render_case_sheet(ws1, cases, statistics)
        
        # Sheet2: 标的信息
        if asset_info:
            ws2 = wb.create_sheet("标的信息")
            self._render_asset_info_sheet(ws2, asset_info)
        
        # Sheet3: 统计分析
        if statistics:
            ws3 = wb.create_sheet("统计分析")
            self._render_statistics_sheet(ws3, statistics)
        
        wb.save(output_path)
        print(f"✅ 估值报告已保存: {output_path}")
    
    def _render_case_sheet(self, ws, cases: List[Dict], statistics: Optional[Dict]):
        """渲染案例工作表"""
        current_row = 1
        
        # 标题
        ws.merge_cells(start_row=current_row, start_column=1,
                      end_row=current_row, end_column=len(self.columns))
        title_cell = ws.cell(row=current_row, column=1, value="不良资产估值 - 参考案例表")
        title_cell.font = Font(name="微软雅黑", size=16, bold=True)
        title_cell.alignment = self.center_alignment
        title_cell.fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
        # 统计信息
        if statistics:
            stats_text = self._format_statistics(statistics)
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(self.columns))
            stats_cell = ws.cell(row=current_row, column=1, value=stats_text)
            stats_cell.font = Font(name="微软雅黑", size=10)
            stats_cell.alignment = self.left_alignment
            ws.row_dimensions[current_row].height = 30
            current_row += 1
        
        current_row += 1
        
        # 表头
        header_row = current_row
        for col_idx, (col_name, width) in enumerate(self.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[header_row].height = 30
        current_row += 1
        
        # 数据
        for case in cases:
            for col_idx, (col_name, width) in enumerate(self.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=case.get(col_name, ""))
                cell.font = self.body_font
                cell.border = self.thin_border
                
                if col_idx in [1, 6, 7]:
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment
            
            # 为参照物位置列添加超链接
            link_url = case.get("参照物位置_链接", "")
            if link_url:
                cell = ws.cell(row=current_row, column=1)
                cell.hyperlink = link_url
                cell.font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
            
            ws.row_dimensions[current_row].height = 25
            current_row += 1
    
    def _render_asset_info_sheet(self, ws, asset_info: Dict):
        """渲染标的信息工作表"""
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        
        ws.cell(row=1, column=1, value="标的资产信息").font = Font(name="微软雅黑", size=14, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.row_dimensions[1].height = 35
        
        row = 3
        for key, value in asset_info.items():
            ws.cell(row=row, column=1, value=key).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=row, column=2, value=value).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
            ws.row_dimensions[row].height = 22
            row += 1
    
    def _render_statistics_sheet(self, ws, statistics: Dict):
        """渲染统计分析工作表"""
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        
        ws.cell(row=1, column=1, value="统计分析").font = Font(name="微软雅黑", size=14, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.row_dimensions[1].height = 35
        
        row = 3
        stat_labels = {
            "case_count": "案例数量",
            "reference_avg_price": "参考均价（元/㎡）",
            "normal_case_count": "有效案例数",
            "min_unit_price": "最低单价（元/㎡）",
            "max_unit_price": "最高单价（元/㎡）",
            "avg_market_value": "平均市值（万元）",
            "min_market_value": "最低市值（万元）",
            "max_market_value": "最高市值（万元）",
        }
        
        for key, label in stat_labels.items():
            if key in statistics:
                ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
                ws.cell(row=row, column=2, value=statistics[key]).font = Font(name="微软雅黑", size=10)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
                ws.row_dimensions[row].height = 22
                row += 1