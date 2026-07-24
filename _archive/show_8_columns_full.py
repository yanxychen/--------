#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""输出完整8列数据（V1格式，完整备注）"""

import glob
import openpyxl

excel_files = sorted(glob.glob('抵押物估值案例_*.xlsx'))
if excel_files:
    latest_file = excel_files[-1]
    wb = openpyxl.load_workbook(latest_file)
    ws = wb.active
    
    print('=' * 100)
    print('📋 不良资产估值 - 参考案例表（V1格式）')
    print('=' * 100)
    
    headers = []
    for col in range(1, 9):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else '')
    
    print('| ' + ' | '.join(headers) + ' |')
    print('|' + '|'.join(['---'] * 8) + '|')
    
    for row in range(2, ws.max_row + 1):
        print('-' * 100)
        print(f'案例{row - 1}:')
        
        row_data = []
        for col in range(1, 9):
            val = ws.cell(row=row, column=col).value
            header = headers[col - 1]
            
            if val is None:
                row_data.append('')
                print(f'  {header}: 空')
            elif isinstance(val, str):
                val = val.replace('\n', ' ')
                row_data.append(val[:50] + '...' if len(val) > 50 else val)
                
                # 备注列显示完整内容
                if header == '备注':
                    print(f'  {header}:')
                    # 按三行格式显示
                    lines = val.split('  ')
                    for i, line in enumerate(lines[:3], 1):
                        print(f'    {line.strip()}')
                else:
                    print(f'  {header}: {val[:60]}{"..." if len(val) > 60 else ""}')
            else:
                row_data.append(str(val))
                print(f'  {header}: {val}')
    
    print('=' * 100)
