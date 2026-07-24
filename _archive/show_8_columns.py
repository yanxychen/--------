#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""输出完整8列数据（V1格式）"""

import glob
import openpyxl

# 读取最新的Excel文件
excel_files = sorted(glob.glob('抵押物估值案例_*.xlsx'))
if excel_files:
    latest_file = excel_files[-1]
    print(f'读取文件: {latest_file}')
    
    wb = openpyxl.load_workbook(latest_file)
    ws = wb.active
    
    print('\n' + '=' * 120)
    print('📋 完整8列数据（V1格式）')
    print('=' * 120)
    
    # 输出表头
    headers = []
    for col in range(1, 9):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else '')
    print('| ' + ' | '.join(headers) + ' |')
    print('|' + '|'.join(['---' * 10] * 8) + '|')
    
    # 输出数据行
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, 9):
            val = ws.cell(row=row, column=col).value
            if val is None:
                row_data.append('')
            elif isinstance(val, str):
                val = val.replace('\n', ' ')
                if len(val) > 40:
                    val = val[:40] + '...'
                row_data.append(val)
            else:
                row_data.append(str(val))
        print('| ' + ' | '.join(row_data) + ' |')
    
    print('=' * 120)
else:
    print('未找到Excel文件')
