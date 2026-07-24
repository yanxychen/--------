#!/usr/bin/env python3
"""
Excel输出质量检查工具
按照8+5项增强标准检查所有Excel文件
"""
import openpyxl
import os
import glob
import re
import requests
from datetime import datetime


class ExcelQualityChecker:
    def __init__(self, check_link_availability=False):
        self.check_link_availability = check_link_availability
        self.expected_headers = [
            "参照物位置",
            "土地面积(m²)",
            "建筑面积(m²)",
            "市场价值(万元)",
            "建筑单价(元/㎡)",
            "数据来源",
            "备注",
            "价格类型"
        ]
        self.valid_statuses = ["正在进行流拍", "已成交", "变卖失败", "正在进行", "即将开始"]
        self.valid_price_types = ["普通司法拍卖"]
        self.auction_stages = ["一拍", "二拍", "三拍", "变卖"]

    def check_file(self, filepath):
        """检查单个Excel文件"""
        result = {
            "filepath": filepath,
            "filename": os.path.basename(filepath),
            "passed": True,
            "issues": [],
            "issue_counts": {
                "header": 0,
                "hyperlink": 0,
                "area_format": 0,
                "remark_format": 0,
                "unit_price": 0,
                "price_type": 0,
                "thousands_separator": 0,
                "enhanced_hyperlink": 0,
                "enhanced_remark_complete": 0,
                "enhanced_sold_price": 0,
                "enhanced_unit_price_logic": 0,
                "enhanced_distance_format": 0
            }
        }

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # 找到数据起始行（跳过标题行/统计行）
            data_start_row = self._find_data_start_row(ws)
            if data_start_row is None:
                result["issues"].append("未找到数据表头行")
                result["passed"] = False
                result["issue_counts"]["header"] += 1
                return result

            # 1. 检查表头
            header_issues = self._check_headers(ws, data_start_row)
            if header_issues:
                result["issues"].extend(header_issues)
                result["issue_counts"]["header"] += len(header_issues)
                result["passed"] = False

            # 获取数据行范围
            max_row = ws.max_row
            data_rows = range(data_start_row + 1, max_row + 1)

            # 逐行检查
            for row_idx in data_rows:
                row_data = self._get_row_data(ws, row_idx)

                # 跳过空行
                if all(v.strip() == '' for v in row_data):
                    continue

                # 跳过统计行/说明行
                first_cell = row_data[0].strip()
                # 数据行应该是：数字+、+内容（如"1、住宅-..."）
                # 排除说明行（如"说明："、"1. 以上案例..."）
                is_data_row = False
                if re.match(r'^\d+、', first_cell):
                    # 以数字+、开头，且后面有实际内容（不是单纯的序号说明）
                    is_data_row = True
                if not is_data_row:
                    continue

                row_num = row_idx - data_start_row

                # 2. 检查超链接
                hyperlink_issue = self._check_hyperlink(ws, row_idx, row_num)
                if hyperlink_issue:
                    result["issues"].append(hyperlink_issue)
                    result["issue_counts"]["hyperlink"] += 1
                    result["passed"] = False

                # 3. 检查面积填写
                area_issues = self._check_area_format(row_data, row_num)
                if area_issues:
                    result["issues"].extend(area_issues)
                    result["issue_counts"]["area_format"] += len(area_issues)
                    result["passed"] = False

                # 4. 检查备注格式
                remark_issues = self._check_remark_format(row_data, row_num)
                if remark_issues:
                    result["issues"].extend(remark_issues)
                    result["issue_counts"]["remark_format"] += len(remark_issues)
                    result["passed"] = False

                # 5. 检查建筑单价
                unit_price_issue = self._check_unit_price(row_data, row_num)
                if unit_price_issue:
                    result["issues"].append(unit_price_issue)
                    result["issue_counts"]["unit_price"] += 1
                    result["passed"] = False

                # 6. 检查价格类型
                price_type_issue = self._check_price_type(row_data, row_num)
                if price_type_issue:
                    result["issues"].append(price_type_issue)
                    result["issue_counts"]["price_type"] += 1
                    result["passed"] = False

                # 7. 检查千分符
                separator_issues = self._check_thousands_separator(row_data, row_num)
                if separator_issues:
                    result["issues"].extend(separator_issues)
                    result["issue_counts"]["thousands_separator"] += len(separator_issues)
                    result["passed"] = False

                # ========== 增强检查 ==========

                # 增强1：超链接有效性检查
                enh_hyperlink_issues = self._check_enhanced_hyperlink(ws, row_idx, row_num)
                if enh_hyperlink_issues:
                    result["issues"].extend(enh_hyperlink_issues)
                    result["issue_counts"]["enhanced_hyperlink"] += len(enh_hyperlink_issues)
                    result["passed"] = False

                # 增强2：备注完整性检查
                enh_remark_issues = self._check_enhanced_remark_complete(row_data, row_num)
                if enh_remark_issues:
                    result["issues"].extend(enh_remark_issues)
                    result["issue_counts"]["enhanced_remark_complete"] += len(enh_remark_issues)
                    result["passed"] = False

                # 增强3：已成交案例价格检查
                enh_sold_issues = self._check_enhanced_sold_price(row_data, row_num)
                if enh_sold_issues:
                    result["issues"].extend(enh_sold_issues)
                    result["issue_counts"]["enhanced_sold_price"] += len(enh_sold_issues)
                    result["passed"] = False

                # 增强4：建筑单价逻辑检查
                enh_unit_price_issues = self._check_enhanced_unit_price_logic(row_data, row_num)
                if enh_unit_price_issues:
                    result["issues"].extend(enh_unit_price_issues)
                    result["issue_counts"]["enhanced_unit_price_logic"] += len(enh_unit_price_issues)
                    result["passed"] = False

                # 增强5：距离格式检查
                enh_distance_issues = self._check_enhanced_distance_format(row_data, row_num)
                if enh_distance_issues:
                    result["issues"].extend(enh_distance_issues)
                    result["issue_counts"]["enhanced_distance_format"] += len(enh_distance_issues)
                    result["passed"] = False

            wb.close()

        except Exception as e:
            result["issues"].append(f"文件读取失败: {str(e)}")
            result["passed"] = False

        return result

    def _find_data_start_row(self, ws):
        """找到数据表头所在行"""
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            first_cell = str(ws.cell(row=row_idx, column=1).value or "").strip()
            if first_cell == "参照物位置" or first_cell.startswith("参照物"):
                return row_idx
        return None

    def _get_row_data(self, ws, row_idx):
        """获取一行数据"""
        data = []
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            data.append(str(cell.value or "").strip())
        return data

    def _check_headers(self, ws, header_row):
        """检查表头"""
        issues = []
        for col_idx, expected in enumerate(self.expected_headers, 1):
            actual = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
            if actual != expected:
                # 放宽比较：去除空格和特殊字符后比较
                actual_clean = re.sub(r'\s+', '', actual)
                expected_clean = re.sub(r'\s+', '', expected)
                if actual_clean != expected_clean:
                    issues.append(f"表头第{col_idx}列不匹配：期望'{expected}'，实际'{actual}'")
        return issues

    def _check_hyperlink(self, ws, row_idx, row_num):
        """检查超链接"""
        cell = ws.cell(row=row_idx, column=1)
        if cell.hyperlink:
            url = cell.hyperlink.target
            # 检查是否指向详情页（不是首页）
            if 'sf-item.taobao.com' in url or 'paimai.jd.com' in url:
                if '/item/' in url or '/sf_item/' in url or 'item.htm' in url:
                    return None
                # 检查是否有具体的item ID
                if re.search(r'(\d{6,})', url):
                    return None
                return f"第{row_num}行超链接可能是首页/列表页: {url}"
            return f"第{row_num}行超链接域名不匹配: {url}"
        # 检查数据来源列是否有链接
        source_cell = ws.cell(row=row_idx, column=6)
        source_url = str(source_cell.value or "").strip()
        if source_url and ('sf-item.taobao.com' in source_url or 'paimai.jd.com' in source_url):
            return f"第{row_num}行参照物位置列无超链接，链接在数据来源列"
        return f"第{row_num}行参照物位置列无超链接"

    def _check_area_format(self, row_data, row_num):
        """检查面积填写格式"""
        issues = []
        land_area = row_data[1]
        building_area = row_data[2]

        # 土地面积：应为"不适用"（无括号）或有效数值
        if land_area:
            if '不适用' in land_area and land_area != '不适用':
                issues.append(f"第{row_num}行土地面积格式错误：'{land_area}'，应为'不适用'（无括号）")
            elif land_area == '0' or land_area == '0.00':
                issues.append(f"第{row_num}行土地面积为0，应填'不适用'")

        # 建筑面积：应为"不适用"（无括号）或有效数值
        if building_area:
            if '不适用' in building_area and building_area != '不适用':
                issues.append(f"第{row_num}行建筑面积格式错误：'{building_area}'，应为'不适用'（无括号）")
            elif building_area == '0' or building_area == '0.00':
                issues.append(f"第{row_num}行建筑面积为0，应填'不适用'")

        return issues

    def _check_remark_format(self, row_data, row_num):
        """检查备注格式"""
        issues = []
        remark = row_data[6]

        if not remark:
            issues.append(f"第{row_num}行备注为空")
            return issues

        # 检查是否包含拍卖轮次
        has_stage = any(stage in remark for stage in self.auction_stages)
        if not has_stage:
            issues.append(f"第{row_num}行备注缺少拍卖轮次（一拍/二拍/三拍/变卖）: {remark[:50]}")

        # 检查是否包含日期（YYYY-MM-DD格式）
        has_date = bool(re.search(r'\d{4}-\d{2}-\d{2}', remark))
        if not has_date:
            issues.append(f"第{row_num}行备注缺少日期(YYYY-MM-DD格式): {remark[:50]}")

        # 检查是否包含价格信息（起拍价/成交价）
        has_start_price = '起拍价' in remark
        has_deal_price = '成交价' in remark

        # 检查价格是否以元为单位且带千分符
        # 查找价格数字
        price_pattern = r'(?:起拍价|成交价)[：:]\s*([\d,]+\.?\d*)\s*(元|万|万元)'
        prices = re.findall(price_pattern, remark)

        if not has_start_price:
            issues.append(f"第{row_num}行备注缺少起拍价: {remark[:50]}")

        # 已成交案例必须有成交价
        is_sold = '已成交' in remark or '成交' in remark
        if is_sold and not has_deal_price:
            issues.append(f"第{row_num}行是已成交案例，但缺少成交价: {remark[:50]}")

        # 检查价格单位是否正确（元为单位）
        for price, unit in prices:
            if unit == '万' or unit == '万元':
                issues.append(f"第{row_num}行价格单位错误，应为'元'不是'万元': {price}{unit}")

        # 检查距离格式
        if '抵押物自身拍卖案例' in remark:
            pass  # 自身拍卖案例的特殊写法
        elif '距离' in remark:
            # 检查距离单位格式
            dist_match = re.search(r'距离[^：:]*[：:]?\s*([\d.]+)\s*(米|公里|km)', remark)
            if dist_match:
                dist = float(dist_match.group(1))
                unit = dist_match.group(2)
                if unit in ['公里', 'km']:
                    # 公里单位：数值应该 >= 1（小于1km应该用米）
                    if dist < 1:
                        issues.append(f"第{row_num}行距离格式错误：小于1km应显示米: {remark[:60]}")
                elif unit == '米':
                    # 米单位：数值应该 < 1000（>=1000米应该用公里）
                    if dist >= 1000:
                        issues.append(f"第{row_num}行距离格式错误：大于等于1km应显示公里: {remark[:60]}")

        # 检查状态
        has_valid_status = any(status in remark for status in self.valid_statuses)
        if not has_valid_status:
            issues.append(f"第{row_num}行备注状态不匹配（应为5种之一）: {remark[:50]}")

        return issues

    def _check_unit_price(self, row_data, row_num):
        """检查建筑单价"""
        building_area = row_data[2]
        unit_price = row_data[4]

        # 如果建筑面积是"不适用"，建筑单价也应为"不适用"
        if building_area == '不适用' or '不适用' in building_area:
            if unit_price != '不适用' and '不适用' not in unit_price:
                # 检查是否是0
                try:
                    val = float(unit_price.replace(',', '').replace('元/㎡', '').strip())
                    if val == 0:
                        return f"第{row_num}行建筑面积不适用，但建筑单价为0，应填'不适用'"
                except:
                    pass

        return None

    def _check_price_type(self, row_data, row_num):
        """检查价格类型"""
        price_type = row_data[7]
        if price_type and price_type not in self.valid_price_types:
            return f"第{row_num}行价格类型错误：'{price_type}'，应为'普通司法拍卖'"
        if not price_type:
            return f"第{row_num}行价格类型为空"
        return None

    def _check_thousands_separator(self, row_data, row_num):
        """检查千分符格式"""
        issues = []
        # 检查市场价值、建筑单价、备注中的价格
        check_cols = {
            3: "市场价值",
            4: "建筑单价"
        }

        for col_idx, col_name in check_cols.items():
            val = row_data[col_idx]
            if val and val != '不适用' and '不适用' not in val:
                # 提取纯数字部分
                num_part = re.sub(r'[^\d,.]', '', val)
                if num_part:
                    # 检查是否有逗号作为千分符
                    try:
                        float_val = float(num_part.replace(',', ''))
                        if float_val >= 1000 and ',' not in num_part:
                            issues.append(f"第{row_num}行{col_name}缺少千分符: {val}")
                    except:
                        pass

        return issues

    def _check_enhanced_hyperlink(self, ws, row_idx, row_num):
        """增强检查1：超链接有效性检查"""
        issues = []
        cell = ws.cell(row=row_idx, column=1)

        if not cell.hyperlink:
            issues.append(f"【增强-超链接】第{row_num}行参照物位置列无超链接")
            return issues

        url = cell.hyperlink.target

        # 检查是否指向详情页，不是首页
        is_homepage = False
        if 'sf.taobao.com' in url and 'sf-item' not in url and 'item' not in url:
            is_homepage = True
        elif 'paimai.jd.com' in url and not re.search(r'/\d+', url):
            is_homepage = True

        if is_homepage:
            issues.append(f"【增强-超链接】第{row_num}行超链接指向首页/列表页，不是详情页: {url}")

        # 检查URL格式是否正确
        if not url.startswith('http'):
            issues.append(f"【增强-超链接】第{row_num}行超链接格式错误: {url}")

        # 检查是否有具体的item ID
        has_item_id = bool(re.search(r'(\d{6,})', url))
        if not has_item_id and 'item' not in url:
            issues.append(f"【增强-超链接】第{row_num}行超链接可能无详情页ID: {url}")

        # 可选：实际检查链接可访问性（耗时，默认关闭）
        if self.check_link_availability:
            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                response = requests.head(url, headers=headers, timeout=5, allow_redirects=True)
                if response.status_code == 404:
                    issues.append(f"【增强-超链接】第{row_num}行超链接404不可访问: {url}")
                elif response.status_code >= 400:
                    issues.append(f"【增强-超链接】第{row_num}行超链接返回状态码{response.status_code}: {url}")
            except Exception as e:
                issues.append(f"【增强-超链接】第{row_num}行超链接访问失败: {str(e)[:50]}")

        return issues

    def _check_enhanced_remark_complete(self, row_data, row_num):
        """增强检查2：备注完整性检查（5要素）"""
        issues = []
        remark = row_data[6]

        if not remark:
            issues.append(f"【增强-备注完整】第{row_num}行备注为空，缺少全部5要素")
            return issues

        required_elements = [
            ("拍卖轮次", ["拍卖轮次", "一拍", "二拍", "三拍", "变卖"]),
            ("时间", ["时间", r'\d{4}-\d{2}-\d{2}']),
            ("价格", ["起拍价"]),
            ("距离", ["距离", "抵押物自身拍卖案例"]),
            ("状态", ["状态：正在进行流拍", "状态：已成交", "状态：变卖失败", "状态：正在进行", "状态：即将开始"] + self.valid_statuses),
        ]

        missing_elements = []
        for elem_name, keywords in required_elements:
            found = False
            for kw in keywords:
                if re.search(kw, remark):
                    found = True
                    break
            if not found:
                missing_elements.append(elem_name)

        if missing_elements:
            issues.append(f"【增强-备注完整】第{row_num}行备注缺少要素: {', '.join(missing_elements)}")

        # 检查每个要素的格式
        # 拍卖轮次格式
        if '拍卖轮次' in remark:
            stage_match = re.search(r'拍卖轮次[：:]\s*(一拍|二拍|三拍|变卖)', remark)
            if not stage_match:
                issues.append(f"【增强-备注完整】第{row_num}行拍卖轮次格式错误，应为'拍卖轮次：一拍/二拍/三拍/变卖'")

        # 时间格式
        if '时间' in remark:
            date_match = re.search(r'时间[：:]\s*(\d{4}-\d{2}-\d{2})', remark)
            if not date_match:
                issues.append(f"【增强-备注完整】第{row_num}行时间格式错误，应为'时间：YYYY-MM-DD'")

        return issues

    def _check_enhanced_sold_price(self, row_data, row_num):
        """增强检查3：已成交案例价格检查"""
        issues = []
        remark = row_data[6]

        if not remark:
            return issues

        # 判断是否是已成交案例
        is_sold = '已成交' in remark
        if not is_sold:
            return issues

        # 检查是否同时有起拍价和成交价
        has_start = '起拍价' in remark
        has_deal = '成交价' in remark

        if not has_start:
            issues.append(f"【增强-已成交价格】第{row_num}行是已成交案例，但缺少起拍价")
        if not has_deal:
            issues.append(f"【增强-已成交价格】第{row_num}行是已成交案例，但缺少成交价")

        # 检查价格单位是否为元，且带千分符
        if has_start:
            start_match = re.search(r'起拍价[：:]\s*([\d,]+\.?\d*)\s*(元|万|万元)', remark)
            if start_match:
                price_str, unit = start_match.group(1), start_match.group(2)
                if unit in ['万', '万元']:
                    issues.append(f"【增强-已成交价格】第{row_num}行起拍价单位错误，应为'元'不是'万元'")
                else:
                    # 检查千分符
                    num_val = float(price_str.replace(',', ''))
                    if num_val >= 1000 and ',' not in price_str:
                        issues.append(f"【增强-已成交价格】第{row_num}行起拍价缺少千分符: {price_str}")

        if has_deal:
            deal_match = re.search(r'成交价[：:]\s*([\d,]+\.?\d*)\s*(元|万|万元)', remark)
            if deal_match:
                price_str, unit = deal_match.group(1), deal_match.group(2)
                if unit in ['万', '万元']:
                    issues.append(f"【增强-已成交价格】第{row_num}行成交价单位错误，应为'元'不是'万元'")
                else:
                    # 检查千分符
                    num_val = float(price_str.replace(',', ''))
                    if num_val >= 1000 and ',' not in price_str:
                        issues.append(f"【增强-已成交价格】第{row_num}行成交价缺少千分符: {price_str}")

        return issues

    def _check_enhanced_unit_price_logic(self, row_data, row_num):
        """增强检查4：建筑单价逻辑检查"""
        issues = []
        building_area_str = row_data[2]
        unit_price_str = row_data[4]

        # 解析建筑面积数值
        building_area_val = None
        building_area_na = False

        if not building_area_str or building_area_str == '不适用' or '不适用' in building_area_str:
            building_area_na = True
        else:
            try:
                clean_val = re.sub(r'[^\d.]', '', building_area_str)
                if clean_val:
                    building_area_val = float(clean_val)
                    if building_area_val == 0:
                        building_area_na = True
            except:
                pass

        # 解析建筑单价数值
        unit_price_val = None
        unit_price_na = False

        if not unit_price_str or unit_price_str == '不适用' or '不适用' in unit_price_str:
            unit_price_na = True
        else:
            try:
                clean_val = re.sub(r'[^\d.]', '', unit_price_str)
                if clean_val:
                    unit_price_val = float(clean_val)
                    if unit_price_val == 0:
                        unit_price_na = True
            except:
                pass

        # 逻辑1：建筑面积为0或不适用 → 建筑单价必须为"不适用"
        if building_area_na:
            if not unit_price_na:
                issues.append(f"【增强-建筑单价逻辑】第{row_num}行建筑面积为不适用/0，但建筑单价不是'不适用': {unit_price_str}")
        # 逻辑2：建筑面积>0 → 建筑单价必须为数值且不能为0
        else:
            if unit_price_na:
                issues.append(f"【增强-建筑单价逻辑】第{row_num}行建筑面积>0，但建筑单价为不适用/0")
            elif unit_price_val is not None and unit_price_val <= 0:
                issues.append(f"【增强-建筑单价逻辑】第{row_num}行建筑面积>0，但建筑单价为0或负数: {unit_price_str}")

        return issues

    def _check_enhanced_distance_format(self, row_data, row_num):
        """增强检查5：距离格式检查"""
        issues = []
        remark = row_data[6]

        if not remark:
            return issues

        # 抵押物自身拍卖案例
        if '抵押物自身拍卖案例' in remark:
            # 检查格式是否正确
            if '距离抵押物：抵押物自身拍卖案例' not in remark:
                issues.append(f"【增强-距离格式】第{row_num}行抵押物自身拍卖案例格式错误，应为'距离抵押物：抵押物自身拍卖案例'")
            return issues

        # 非自身案例，检查距离格式
        dist_match = re.search(r'距离抵押物[：:]\s*([\d.]+)\s*(米|公里)', remark)
        if not dist_match:
            # 看看有没有距离相关内容
            if '距离' in remark:
                issues.append(f"【增强-距离格式】第{row_num}行距离格式无法识别")
            return issues

        dist = float(dist_match.group(1))
        unit = dist_match.group(2)

        if unit == '米':
            # 米：应为整数，且 < 1000
            if dist >= 1000:
                issues.append(f"【增强-距离格式】第{row_num}行距离>=1000米，应改用公里单位: {dist}米")
            elif dist != int(dist):
                issues.append(f"【增强-距离格式】第{row_num}行米单位距离不应有小数: {dist}米，应为{int(dist)}米")
        elif unit == '公里':
            # 公里：应 >= 1，且保留一位小数
            if dist < 1:
                issues.append(f"【增强-距离格式】第{row_num}行距离<1公里，应改用米单位: {dist}公里")
            else:
                # 检查是否有一位小数
                if '.' in str(dist):
                    decimal_part = str(dist).split('.')[1]
                    if len(decimal_part) != 1:
                        issues.append(f"【增强-距离格式】第{row_num}行公里单位应保留一位小数: {dist}公里")

        return issues

    def check_all_files(self, directories):
        """检查所有目录下的Excel文件"""
        all_files = []
        for directory in directories:
            if os.path.exists(directory):
                pattern = os.path.join(directory, '*.xlsx')
                files = glob.glob(pattern)
                # 排除临时文件（~开头）
                files = [f for f in files if not os.path.basename(f).startswith('~')]
                all_files.extend(files)

        all_files.sort()

        results = []
        for filepath in all_files:
            result = self.check_file(filepath)
            results.append(result)

        return results

    def generate_report(self, results):
        """生成检查报告"""
        total = len(results)
        passed = sum(1 for r in results if r["passed"])
        failed = total - passed

        # 统计各类问题（基础）
        issue_totals = {
            "header": 0,
            "hyperlink": 0,
            "area_format": 0,
            "remark_format": 0,
            "unit_price": 0,
            "price_type": 0,
            "thousands_separator": 0,
            "enhanced_hyperlink": 0,
            "enhanced_remark_complete": 0,
            "enhanced_sold_price": 0,
            "enhanced_unit_price_logic": 0,
            "enhanced_distance_format": 0
        }

        for r in results:
            for key in issue_totals:
                issue_totals[key] += r["issue_counts"].get(key, 0)

        # 统计增强检查的文件覆盖数
        enh_files = {
            "enhanced_hyperlink": 0,
            "enhanced_remark_complete": 0,
            "enhanced_sold_price": 0,
            "enhanced_unit_price_logic": 0,
            "enhanced_distance_format": 0
        }
        for r in results:
            for key in enh_files:
                if r["issue_counts"].get(key, 0) > 0:
                    enh_files[key] += 1

        report = []
        report.append("=" * 70)
        report.append("Excel输出质量增强检查报告")
        report.append("=" * 70)
        report.append(f"检查时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"总文件数: {total}")
        report.append(f"合格文件数: {passed}")
        report.append(f"不合格文件数: {failed}")
        if total > 0:
            report.append(f"合格率: {passed/total*100:.1f}%")
        report.append("")

        report.append("【基础检查问题统计】")
        report.append(f"  - 表头问题: {issue_totals['header']} 处")
        report.append(f"  - 超链接问题: {issue_totals['hyperlink']} 处")
        report.append(f"  - 面积填写格式错误: {issue_totals['area_format']} 处")
        report.append(f"  - 备注格式问题: {issue_totals['remark_format']} 处")
        report.append(f"  - 建筑单价问题: {issue_totals['unit_price']} 处")
        report.append(f"  - 价格类型错误: {issue_totals['price_type']} 处")
        report.append(f"  - 数值缺少千分符: {issue_totals['thousands_separator']} 处")
        report.append("")

        report.append("【增强检查问题统计】")
        report.append(f"  1. 超链接有效性: {enh_files['enhanced_hyperlink']} 个文件，{issue_totals['enhanced_hyperlink']} 个问题")
        report.append(f"  2. 备注完整性(5要素): {enh_files['enhanced_remark_complete']} 个文件，{issue_totals['enhanced_remark_complete']} 个问题")
        report.append(f"  3. 已成交案例价格缺失: {enh_files['enhanced_sold_price']} 个文件，{issue_totals['enhanced_sold_price']} 个问题")
        report.append(f"  4. 建筑单价逻辑错误: {enh_files['enhanced_unit_price_logic']} 个文件，{issue_totals['enhanced_unit_price_logic']} 个问题")
        report.append(f"  5. 距离格式错误: {enh_files['enhanced_distance_format']} 个文件，{issue_totals['enhanced_distance_format']} 个问题")
        report.append("")

        report.append("-" * 70)
        report.append("详细问题列表:")
        report.append("-" * 70)

        for r in results:
            status = "✓ 合格" if r["passed"] else "✗ 不合格"
            report.append(f"\n【{status}】{r['filename']}")
            if r["issues"]:
                # 分类显示
                basic_issues = [i for i in r["issues"] if not i.startswith("【增强-")]
                enh_issues = [i for i in r["issues"] if i.startswith("【增强-")]

                if basic_issues:
                    report.append("  基础问题:")
                    for issue in basic_issues[:5]:
                        report.append(f"    · {issue}")
                    if len(basic_issues) > 5:
                        report.append(f"    ... 还有 {len(basic_issues)-5} 个基础问题")

                if enh_issues:
                    report.append("  增强问题:")
                    for issue in enh_issues[:10]:
                        report.append(f"    · {issue}")
                    if len(enh_issues) > 10:
                        report.append(f"    ... 还有 {len(enh_issues)-10} 个增强问题")

        report.append("")
        report.append("=" * 70)
        return "\n".join(report)


def main():
    checker = ExcelQualityChecker()

    # 检查的目录
    directories = [
        '/workspace',
        '/workspace/temp_files',
        '/workspace/test_results'
    ]

    print("开始检查Excel文件质量...")
    print(f"检查目录: {directories}")
    print()

    results = checker.check_all_files(directories)
    report = checker.generate_report(results)

    print(report)

    # 保存报告
    report_file = '/workspace/quality_report.txt'
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"\n报告已保存到: {report_file}")


if __name__ == '__main__':
    main()
